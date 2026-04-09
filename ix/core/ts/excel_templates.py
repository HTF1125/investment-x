"""Excel workbook generation for timeseries export and download templates."""

from __future__ import annotations

import io
from typing import Dict

import pandas as pd

from .mutations import BULK_EXAMPLE, BULK_META_FIELDS

# Universal Excel formula template for the download endpoint.
DOWNLOAD_FORMULA_TEMPLATE = (
    """=IF(__C__6="FactSet",IF(ISNUMBER(SEARCH(__C__3, "FDS_ECON_DATA; FDS_COM_DATA",1)),"""
    """FDSC("","","PSETCAL(SEVENDAY);"&__C__3&"('" & __C__2 & "'," & $B$1 & "," & $C$1 & ", D, NONE, NONE)"),\n"""
    """FDSC("","","PSETCAL(SEVENDAY);NO_REPEAT_F(SPEC_ID_DATA('" & __C__2 & ":" & __C__3 & "','" & $B$1 & "','" & $C$1 & "', D, NONE, NONE,2))")),\n"""
    """IF(__C__6="Bloomberg",BDH(__C__2,__C__3,$B$1,$C$1,"SORT", "FALSE","DTS", "FALSE","DAYS", "C","FILL", "B"),\n"""
    """IF(__C__6="Infomax",IMDH(__C__5,__C__2&"",__C__3,$B$1+2,$C$1,9999,"Per=\uc77c,sort=A,real=false,Bizday=12,Quote=\uc885\uac00,ROUND=9,Pos=20,Orient=V,Title="&__C__7&",DtFmt=1,TmFmt=1,unit=true"),\n"""
    """NA())))"""
)


def generate_export_workbook(all_ts: list) -> io.BytesIO:
    """Build an Excel workbook with all timeseries metadata.

    Returns a BytesIO buffer containing the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    header_font = Font(name="Consolas", size=9, bold=True)
    header_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    normal_font = Font(name="Consolas", size=9)

    # Header row
    for ci, field in enumerate(BULK_META_FIELDS):
        cell = ws.cell(row=1, column=ci + 1, value=field)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for ri, ts in enumerate(all_ts, start=2):
        for ci, field in enumerate(BULK_META_FIELDS):
            val = getattr(ts, field, None)
            cell = ws.cell(row=ri, column=ci + 1, value=val if val is not None else "")
            cell.font = normal_font

    # Auto-size columns
    widths = {"code": 18, "name": 26, "source_code": 24, "remark": 28}
    for ci, field in enumerate(BULK_META_FIELDS):
        ws.column_dimensions[get_column_letter(ci + 1)].width = widths.get(field, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_create_template_workbook() -> io.BytesIO:
    """Build a blank Excel template for bulk timeseries creation.

    Sheet 1 "Metadata": one row per timeseries (columns = fields).
    Sheet 2 "Data": columnar time-series data (Date | code1 | code2 | ...).

    Returns a BytesIO buffer containing the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # -- Sheet 1: Metadata --
    ws_meta = wb.active
    ws_meta.title = "Metadata"

    header_font = Font(name="Consolas", size=9, bold=True)
    required_font = Font(name="Consolas", size=9, bold=True, color="CC0000")
    example_font = Font(name="Consolas", size=9, color="999999", italic=True)
    header_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    # Header row
    for ci, field in enumerate(BULK_META_FIELDS):
        label = f"{field} *" if field == "code" else field
        cell = ws_meta.cell(row=1, column=ci + 1, value=label)
        cell.font = required_font if field == "code" else header_font
        cell.fill = header_fill

    # Example row (row 2)
    for ci, field in enumerate(BULK_META_FIELDS):
        val = BULK_EXAMPLE.get(field, "")
        cell = ws_meta.cell(row=2, column=ci + 1, value=val)
        cell.font = example_font

    # Column widths
    widths = {"code": 18, "name": 22, "source_code": 22, "remark": 28}
    for ci, field in enumerate(BULK_META_FIELDS):
        ws_meta.column_dimensions[get_column_letter(ci + 1)].width = widths.get(
            field, 14
        )

    # -- Sheet 2: Data --
    ws_data = wb.create_sheet(title="Data")

    ws_data.cell(row=1, column=1, value="Date").font = header_font
    ws_data.cell(row=1, column=1).fill = header_fill
    # Example code header
    ws_data.cell(row=1, column=2, value="US_CPI_YOY").font = example_font
    ws_data.cell(row=1, column=2).fill = header_fill

    # Example dates + values
    example_dates = pd.date_range(
        end=pd.Timestamp.now().normalize(), periods=5, freq="ME"
    )
    for i, dt in enumerate(example_dates):
        cell = ws_data.cell(row=2 + i, column=1, value=dt.to_pydatetime())
        cell.font = example_font
        cell.number_format = "YYYY-MM-DD"
        ws_data.cell(
            row=2 + i, column=2, value=round(2.0 + i * 0.3, 1)
        ).font = example_font

    ws_data.column_dimensions["A"].width = 14
    ws_data.column_dimensions["B"].width = 16

    # -- Save --
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_download_template_workbook(
    grouped: Dict[str, list],
    start_dt: pd.Timestamp,
    end_dt: pd.Timestamp,
) -> io.BytesIO:
    """Build the download-template Excel workbook.

    *grouped* maps source name to a list of Timeseries objects.
    Returns a BytesIO buffer containing the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    start_str = start_dt.strftime("%Y-%m-%d")
    end_str = end_dt.strftime("%Y-%m-%d")

    wb = Workbook()
    wb.remove(wb.active)

    # Styles -- white background, black text
    header_font = Font(name="Consolas", size=9, color="000000")
    value_font = Font(name="Consolas", size=9, color="000000")
    code_font = Font(name="Consolas", size=9, color="000000", bold=True)
    formula_font = Font(name="Consolas", size=8, color="000000", italic=True)
    date_font = Font(name="Consolas", size=9, color="000000")

    # Date range extent (for formula row count)
    date_top = end_dt

    for sheet_source, ts_list in grouped.items():
        ws = wb.create_sheet(title=(sheet_source or "Unknown")[:31])

        # Row 1: Header info
        ws.cell(row=1, column=1, value=sheet_source).font = Font(
            name="Consolas", size=10, color="0000FF", bold=True
        )
        ws.cell(row=1, column=2, value=start_dt.to_pydatetime()).font = value_font
        ws.cell(row=1, column=2).number_format = "mm-dd-yy"
        ws.cell(row=1, column=3, value=end_dt.to_pydatetime()).font = value_font
        ws.cell(row=1, column=3).number_format = "mm-dd-yy"

        # Metadata row labels (col A)
        row_labels = {
            2: "source_ticker",
            3: "source_field",
            4: "source_code",
            5: "asset_class",
            6: "source",
            7: "name",
            8: "code",
        }
        for r, label in row_labels.items():
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = header_font

        # Fill columns B+ with timeseries metadata
        for col_idx, ts in enumerate(ts_list, start=2):
            sc = str(ts.source_code) if ts.source_code else ":"
            parts = sc.rsplit(":", 1) if ":" in sc else [sc, ""]
            ticker = parts[0] if len(parts) > 0 else ""
            field = parts[1] if len(parts) > 1 else ""

            c = ws.cell(row=2, column=col_idx, value=ticker)
            c.font = value_font
            c.number_format = "@"
            c.data_type = "s"
            c = ws.cell(row=3, column=col_idx, value=field)
            c.font = value_font
            c.number_format = "@"
            c.data_type = "s"
            c = ws.cell(row=4, column=col_idx, value=sc)
            c.font = value_font
            c.number_format = "@"
            c.data_type = "s"
            ws.cell(
                row=5, column=col_idx, value=ts.asset_class or ""
            ).font = value_font
            ws.cell(row=6, column=col_idx, value=ts.source or "").font = value_font
            ws.cell(row=7, column=col_idx, value=ts.name or "").font = value_font
            c = ws.cell(row=8, column=col_idx, value=ts.code or "")
            c.font = code_font
            c.number_format = "@"
            c.data_type = "s"

            # Universal formula -- column letter swapped via template
            col = get_column_letter(col_idx)
            formula_text = DOWNLOAD_FORMULA_TEMPLATE.replace("__C__", col)

            ws.cell(row=9, column=col_idx, value=formula_text).font = formula_font

        # Date column (A9+) -- ascending from start_date toward end_date+2
        num_date_rows = (date_top - start_dt).days + 1
        c = ws.cell(row=9, column=1, value="=B1")
        c.font = date_font
        c.number_format = "mm-dd-yy"
        for i in range(1, num_date_rows):
            c = ws.cell(row=9 + i, column=1, value=f"=A{9 + i - 1}+1")
            c.font = date_font
            c.number_format = "mm-dd-yy"

        # Column widths
        ws.column_dimensions["A"].width = 16
        for col_idx in range(2, len(ts_list) + 2):
            ws.column_dimensions[
                ws.cell(row=2, column=col_idx).column_letter
            ].width = 14

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
