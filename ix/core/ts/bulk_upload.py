"""Bulk timeseries creation and template upload processing."""

from __future__ import annotations

import io
import math
import time as _time
from datetime import date as _date
from datetime import datetime, datetime as _dt
from typing import Optional

import pandas as pd
from sqlalchemy.orm import Session as SessionType

from ix.common import get_logger
from ix.db.conn import Session, ensure_connection
from ix.db.models import Timeseries
from .mutations import BULK_META_FIELDS, apply_timeseries_updates

logger = get_logger(__name__)


def _parse_date_cell(date_cell) -> Optional[str]:
    """Convert an Excel cell value to a 'YYYY-MM-DD' string, or None."""
    if isinstance(date_cell, (_dt, _date)):
        return date_cell.strftime("%Y-%m-%d")
    if isinstance(date_cell, (int, float)):
        try:
            return pd.to_datetime(
                date_cell, unit="D", origin="1899-12-30"
            ).strftime("%Y-%m-%d")
        except Exception:
            return None
    if isinstance(date_cell, str):
        try:
            return pd.to_datetime(date_cell).strftime("%Y-%m-%d")
        except Exception:
            return None
    return None


def process_bulk_create(contents: bytes) -> dict:
    """Parse bulk-create template: create timeseries metadata + merge data.

    Raises ValueError for bad input (caller should map to HTTP 400).
    """
    from openpyxl import load_workbook

    _t0 = _time.time()
    logger.info("Bulk create: parsing %d bytes...", len(contents))
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Invalid Excel format: {e}")

    # -- Parse Metadata sheet --
    ws_meta = wb["Metadata"] if "Metadata" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws_meta.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError(
            "Metadata sheet must have a header row and at least one data row."
        )

    # Map header to column index
    headers = [
        str(h).strip().replace(" *", "").replace("*", "") if h else ""
        for h in rows[0]
    ]
    field_indices = {}
    for fi, field in enumerate(BULK_META_FIELDS):
        if field in headers:
            field_indices[field] = headers.index(field)

    if "code" not in field_indices:
        raise ValueError("Metadata sheet must contain a 'code' column.")

    ts_metas: list[dict] = []
    for row in rows[1:]:
        code_idx = field_indices["code"]
        code_val = row[code_idx] if code_idx < len(row) else None
        if not code_val or not str(code_val).strip():
            continue
        meta: dict = {}
        for field, idx in field_indices.items():
            val = row[idx] if idx < len(row) else None
            if val is not None and str(val).strip():
                meta[field] = str(val).strip()
        ts_metas.append(meta)

    if not ts_metas:
        raise ValueError("No timeseries found in Metadata sheet.")

    logger.info("Bulk create: found %d timeseries definitions", len(ts_metas))

    # -- Create / update timeseries in DB --
    ensure_connection()
    created_codes: list[str] = []
    updated_codes: list[str] = []
    errors: list[str] = []

    with Session() as db:
        codes = [m["code"] for m in ts_metas]
        existing = db.query(Timeseries).filter(Timeseries.code.in_(codes)).all()
        existing_by_code = {ts.code: ts for ts in existing}

        for meta in ts_metas:
            code = meta["code"]
            try:
                ts = existing_by_code.get(code)
                if ts is None:
                    ts = Timeseries(code=code)
                    db.add(ts)
                    existing_by_code[code] = ts
                    created_codes.append(code)
                else:
                    updated_codes.append(code)

                apply_timeseries_updates(ts, meta)
                db.flush()
            except Exception as e:
                logger.error("Bulk create error for %s: %s", code, e)
                errors.append(f"{code}: {e}")

        db.commit()

    logger.info(
        "Bulk create: %d created, %d updated in %.1fs",
        len(created_codes),
        len(updated_codes),
        _time.time() - _t0,
    )

    # -- Parse Data sheet (optional) --
    data_result = {}
    if "Data" in wb.sheetnames:
        ws_data = wb["Data"]
        data_rows = list(ws_data.iter_rows(values_only=True))

        if len(data_rows) >= 2:
            # Row 0 = headers: Date, code1, code2, ...
            data_headers = data_rows[0]
            data_codes = [
                str(h).strip() for h in data_headers[1:] if h and str(h).strip()
            ]

            if data_codes:
                all_records: dict[str, dict] = {c: {} for c in data_codes}

                for row in data_rows[1:]:
                    date_cell = row[0] if row else None
                    if date_cell is None:
                        continue
                    date_str = _parse_date_cell(date_cell)
                    if not date_str:
                        continue

                    for ci, code in enumerate(data_codes):
                        col_idx = ci + 1
                        if col_idx < len(row) and row[col_idx] is not None:
                            try:
                                val = float(row[col_idx])
                                if not math.isnan(val):
                                    all_records[code][date_str] = val
                            except (ValueError, TypeError):
                                continue

                # Remove empty codes
                all_records = {c: d for c, d in all_records.items() if d}

                if all_records:
                    all_dates = sorted(
                        {d for rec in all_records.values() for d in rec}
                    )
                    columns = {
                        code: [rec.get(d) for d in all_dates]
                        for code, rec in all_records.items()
                    }
                    dates_index = pd.to_datetime(all_dates, errors="coerce")
                    df = pd.DataFrame(columns, index=dates_index)
                    df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)

                    if not df.empty:
                        with Session() as local_db:
                            data_result = merge_columnar_to_db(df, local_db)
                        logger.info(
                            "Bulk create: merged %d data points for %d codes",
                            data_result.get("points", 0),
                            len(data_result.get("updated", [])),
                        )

    wb.close()

    response = {
        "message": f"Created {len(created_codes)}, updated {len(updated_codes)} timeseries.",
        "created": created_codes,
        "updated": updated_codes,
    }
    if errors:
        response["errors"] = errors
    if data_result:
        response["data_merged"] = data_result.get("points", 0)
        response["data_codes"] = data_result.get("updated", [])
        if data_result.get("not_found"):
            response["data_not_found"] = data_result["not_found"]

    return response


def process_template_upload(contents: bytes) -> dict:
    """Parse filled Excel download template and merge data into DB.

    Raises ValueError for bad input (caller should map to HTTP 400).
    """
    from openpyxl import load_workbook

    _t0 = _time.time()
    logger.info("Template upload: parsing %d bytes...", len(contents))
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Invalid Excel format: {e}")
    logger.info("Template upload: workbook loaded in %.1fs", _time.time() - _t0)

    all_records: dict = {}  # {code: {date_str: value}}
    for ws in wb.worksheets:
        row_num = 0
        codes_by_col_idx: dict = {}  # {0-based col index: code}
        for row in ws.iter_rows(values_only=True):
            row_num += 1
            if row_num < 8:
                continue
            if row_num == 8:
                # Row 8 = codes (skip col A at index 0)
                for ci, val in enumerate(row):
                    if ci > 0 and val:
                        codes_by_col_idx[ci] = str(val).strip()
                continue
            # Rows 9+ = data
            if not row or not codes_by_col_idx:
                continue
            date_cell = row[0]
            if date_cell is None:
                continue
            date_str = _parse_date_cell(date_cell)
            if not date_str:
                continue
            for ci, code in codes_by_col_idx.items():
                if ci < len(row) and row[ci] is not None:
                    try:
                        val = float(row[ci])
                        if math.isnan(val):
                            continue
                    except Exception:
                        continue
                    if code not in all_records:
                        all_records[code] = {}
                    all_records[code][date_str] = val
    wb.close()
    logger.info(
        "Template upload: parsed %d codes in %.1fs",
        len(all_records),
        _time.time() - _t0,
    )

    if not all_records:
        raise ValueError("No data found in uploaded file.")

    # Build columnar format
    all_dates = sorted({d for rec in all_records.values() for d in rec})
    columns = {}
    for code, rec in all_records.items():
        columns[code] = [rec.get(d) for d in all_dates]

    # Merge into local DB (fresh session for thread safety)
    ensure_connection()
    dates_index = pd.to_datetime(all_dates, errors="coerce")
    df = pd.DataFrame(columns, index=dates_index)
    df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)
    if df.empty:
        raise ValueError("No valid records after cleaning.")

    logger.info("Template upload: merging %d codes into local DB...", len(df.columns))
    _t1 = _time.time()
    with Session() as local_db:
        result = merge_columnar_to_db(df, local_db)
    logger.info(
        "Template upload: local DB merge done in %.1fs (%d updated)",
        _time.time() - _t1,
        len(result["updated"]),
    )
    response = {
        "message": f"Merged {result['points']} points for {len(result['updated'])} codes.",
        "db_updated": result["updated"],
        "db_points_merged": result["points"],
    }
    if result["not_found"]:
        response["warning"] = f"Codes not found in database: {result['not_found']}"

    logger.info(
        "Template upload: returning response after %.1fs", _time.time() - _t0
    )
    return response


def merge_columnar_to_db(df: pd.DataFrame, db: SessionType) -> dict:
    """Merge a date-indexed DataFrame (columns=codes) into the database.

    Returns {"updated": [...], "not_found": [...], "points": int}.
    Uses batch loading to minimise DB queries.
    """
    from ix.db.models import TimeseriesData

    codes_list = list(df.columns)
    if not codes_list:
        return {"updated": [], "not_found": [], "points": 0}

    # Batch-load all matching Timeseries in one query
    all_ts = db.query(Timeseries).filter(Timeseries.code.in_(codes_list)).all()
    ts_by_code = {ts.code: ts for ts in all_ts}
    found_codes = set(ts_by_code.keys())
    not_found_codes = [c for c in codes_list if c not in found_codes]

    if not ts_by_code:
        return {"updated": [], "not_found": not_found_codes, "points": 0}

    # Batch-load all data records in one query
    ts_ids = [ts.id for ts in all_ts]
    all_data = (
        db.query(TimeseriesData)
        .filter(TimeseriesData.timeseries_id.in_(ts_ids))
        .all()
    )
    data_by_ts_id = {dr.timeseries_id: dr for dr in all_data}

    updated_codes = []
    total_points = 0
    now = datetime.now()

    for code in codes_list:
        ts = ts_by_code.get(code)
        if ts is None:
            continue

        # Get or create data record (no extra query -- already loaded)
        data_record = data_by_ts_id.get(ts.id)
        if data_record is None:
            data_record = TimeseriesData(timeseries_id=ts.id, data={})
            db.add(data_record)
            data_by_ts_id[ts.id] = data_record

        column_data = data_record.data if data_record.data else {}

        if column_data and isinstance(column_data, dict):
            existing_data = pd.Series(column_data)
            if not existing_data.empty:
                existing_data.index = pd.to_datetime(
                    existing_data.index, errors="coerce"
                )
                existing_data = existing_data.dropna()
        else:
            existing_data = pd.Series(dtype=float)

        new_series = df[code].dropna()

        if not existing_data.empty:
            combined = pd.concat([existing_data, new_series], axis=0)
            combined = combined[~combined.index.duplicated(keep="last")].sort_index()
        else:
            combined = new_series.sort_index()

        data_dict = {}
        for k, v in combined.to_dict().items():
            date_str = (
                str(k.date())
                if hasattr(k, "date")
                else str(pd.to_datetime(k).date())
                if not isinstance(k, str)
                else k
            )
            data_dict[date_str] = (
                float(v)
                if v is not None and not (isinstance(v, float) and pd.isna(v))
                else None
            )

        data_record.data = data_dict
        data_record.updated = now
        ts.start = combined.index.min().date() if len(combined) > 0 else None
        ts.end = combined.index.max().date() if len(combined) > 0 else None
        ts.num_data = len(combined)
        ts.updated = now

        total_points += len(new_series)
        updated_codes.append(code)

    db.commit()
    return {
        "updated": updated_codes,
        "not_found": not_found_codes,
        "points": total_points,
    }
