"""
Timeseries router for timeseries data management.
"""

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    Request,
    status,
    UploadFile,
    File,
    BackgroundTasks,
)
from fastapi.responses import Response, StreamingResponse
from typing import Optional, List, Dict, Any
from collections import OrderedDict
import json
import math
import pandas as pd
from datetime import datetime, date

from ix.api.schemas import (
    TimeseriesResponse,
    TimeseriesCreate,
    TimeseriesBulkUpdate,
    TimeseriesDataUpload,
    TimeseriesColumnarUpload,
    TimeseriesUpdate,
)
from ix.api.dependencies import get_db, get_current_admin_user
from ix.db.models import Timeseries
from ix.db.conn import ensure_connection, Session
from ix.db.models.user import User
from ix.db.query import *
from sqlalchemy.orm import joinedload, Session as SessionType
from ix.misc import get_logger

logger = get_logger(__name__)

router = APIRouter()


def _apply_timeseries_updates(ts: Timeseries, data: dict) -> None:
    """Apply whitelisted updates to a Timeseries instance."""
    if "code" in data and data["code"] is not None:
        ts.code = str(data["code"])

    if "name" in data:
        name = data["name"]
        ts.name = str(name)[:200] if name else None
    if "provider" in data:
        provider = data["provider"]
        ts.provider = str(provider)[:100] if provider else None
    if "asset_class" in data:
        asset_class = data["asset_class"]
        ts.asset_class = str(asset_class)[:50] if asset_class else None
    if "category" in data:
        category = data["category"]
        ts.category = str(category)[:100] if category else None
    if "source" in data:
        source = data["source"]
        ts.source = str(source)[:100] if source else None
    if "source_code" in data:
        source_code = data["source_code"]
        ts.source_code = str(source_code)[:2000] if source_code else None
    if "frequency" in data:
        frequency = data["frequency"]
        ts.frequency = str(frequency)[:20] if frequency else None
    if "unit" in data:
        unit = data["unit"]
        ts.unit = str(unit)[:50] if unit else None
    if "scale" in data:
        scale = data["scale"]
        if scale is not None:
            ts.scale = int(scale)
        else:
            ts.scale = None
    if "currency" in data:
        currency = data["currency"]
        ts.currency = str(currency)[:10] if currency else None
    if "country" in data:
        country = data["country"]
        ts.country = str(country)[:100] if country else None
    if "remark" in data:
        remark = data["remark"]
        ts.remark = str(remark) if remark else None
    if "favorite" in data:
        ts.favorite = bool(data["favorite"]) if data["favorite"] is not None else None


@router.get("/timeseries", response_model=List[TimeseriesResponse])
async def get_timeseries(
    limit: Optional[int] = Query(None, description="Limit number of results"),
    offset: int = Query(0, description="Offset for pagination"),
    search: Optional[str] = Query(
        None, description="Search by code, name, source, or category"
    ),
    category: Optional[str] = Query(None, description="Filter by category"),
    asset_class: Optional[str] = Query(None, description="Filter by asset class"),
    provider: Optional[str] = Query(None, description="Filter by provider"),
    db: SessionType = Depends(get_db),
):
    """
    GET /api/timeseries - List all timeseries with optional filtering and pagination.
    """
    ensure_connection()
    from sqlalchemy import or_

    timeseries_query = db.query(Timeseries)

    # Full-text search across multiple fields
    if search:
        pattern = f"%{search}%"
        timeseries_query = timeseries_query.filter(
            or_(
                Timeseries.code.ilike(pattern),
                Timeseries.name.ilike(pattern),
                Timeseries.source.ilike(pattern),
                Timeseries.category.ilike(pattern),
                Timeseries.source_code.ilike(pattern),
            )
        )

    # Apply filters
    if category:
        timeseries_query = timeseries_query.filter(Timeseries.category == category)
    if asset_class:
        timeseries_query = timeseries_query.filter(
            Timeseries.asset_class == asset_class
        )
    if provider:
        timeseries_query = timeseries_query.filter(Timeseries.provider == provider)

    # Apply pagination
    if offset:
        timeseries_query = timeseries_query.offset(offset)
    if limit:
        timeseries_query = timeseries_query.limit(limit)

    timeseries_list = timeseries_query.all()

    formatted_timeseries = []
    for ts in timeseries_list:
        formatted_ts = TimeseriesResponse(
            id=str(ts.id),
            code=str(ts.code) if ts.code else None,
            name=str(ts.name) if ts.name else None,
            provider=ts.provider,
            asset_class=str(ts.asset_class) if ts.asset_class else None,
            category=ts.category,
            start=ts.start,
            end=ts.end,
            num_data=int(ts.num_data) if ts.num_data is not None else None,
            source=ts.source,
            source_code=getattr(ts, "source_code", None),
            frequency=str(ts.frequency) if ts.frequency else None,
            unit=getattr(ts, "unit", None),
            scale=getattr(ts, "scale", None),
            currency=getattr(ts, "currency", None),
            country=getattr(ts, "country", None),
            remark=str(ts.remark) if ts.remark else None,
            favorite=bool(ts.favorite) if ts.favorite is not None else False,
        )
        formatted_timeseries.append(formatted_ts)

    logger.info(f"Retrieved {len(formatted_timeseries)} timeseries records")
    return formatted_timeseries


@router.get("/timeseries/sources")
async def get_timeseries_sources(
    db: SessionType = Depends(get_db),
):
    """Return distinct source names for timeseries that have source_code set."""
    from sqlalchemy import distinct

    ensure_connection()
    rows = (
        db.query(distinct(Timeseries.source))
        .filter(Timeseries.source_code.isnot(None))
        .filter(Timeseries.source.isnot(None))
        .all()
    )
    return sorted([r[0] for r in rows])


# Universal Excel formula template for the download endpoint.
# Matches user's EXACT formatted string (including spaces).
# __C__ is replaced with the actual column letter (B, C, D, ...) per timeseries.
_DOWNLOAD_FORMULA_TEMPLATE = (
    """=IF(__C__6="FactSet",IF(ISNUMBER(SEARCH(__C__3, "FDS_ECON_DATA; FDS_COM_DATA",1)),"""
    """FDSC("","","PSETCAL(SEVENDAY);"&__C__3&"('" & __C__2 & "'," & $D$1+2 & "," & $C$1 & ", D, NONE, NONE)"),\n"""
    """FDSC("","","PSETCAL(SEVENDAY);NO_REPEAT_F(SPEC_ID_DATA('" & __C__2 & ":" & __C__3 & "','" & $D$1+2 & "','" & $C$1 & "', D, NONE, NONE,2))")),\n"""
    """IF(__C__6="Bloomberg",BDH(__C__2,__C__3,$C$1,$D$1+2,"SORT", "TRUE","DTS", "FALSE","DAYS", "C","FILL", "B"),\n"""
    """IF(__C__6="Infomax",IMDH(__C__5,__C__2&"",__C__3,$C$1,$D$1,9999,"Per=일,sort=D,real=false,Bizday=12,Quote=종가,ROUND=9,Pos=20,Orient=V,Title="&__C__7&",DtFmt=1,TmFmt=1,unit=true"),\n"""
    """NA())))"""
)


@router.get("/timeseries/download_template")
async def download_template(
    source: Optional[List[str]] = Query(
        None,
        description="Filter by source(s): Bloomberg, FactSet, Infomax (repeatable)",
    ),
    start_date: Optional[str] = Query(
        None, description="Start date (ISO-8601), default ~2 years ago"
    ),
    end_date: Optional[str] = Query(
        None, description="End date (ISO-8601), default today"
    ),
    db: SessionType = Depends(get_db),
):
    """Generate an Excel template for data download/upload workflow.

    Produces a .xlsx file with metadata rows and pre-built formulas
    matching the Google Sheets Bloomberg/FactSet/Infomax workflow.
    Accepts multiple source values via repeated query params.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    ensure_connection()

    # Date range
    end_dt = pd.to_datetime(end_date) if end_date else pd.Timestamp.now()
    start_dt = (
        pd.to_datetime(start_date) if start_date else end_dt - pd.DateOffset(years=2)
    )
    start_str = start_dt.strftime("%Y-%m-%d")
    end_str = end_dt.strftime("%Y-%m-%d")

    # Query timeseries that have source_code set
    query = db.query(Timeseries).filter(Timeseries.source_code.isnot(None))
    if source:
        query = query.filter(Timeseries.source.in_(source))
    all_ts = query.order_by(Timeseries.source, Timeseries.code).all()

    if not all_ts:
        raise HTTPException(
            status_code=404, detail="No timeseries with source_code found."
        )

    # Group by source
    grouped: Dict[str, list] = {}
    for ts in all_ts:
        src = ts.source or "Unknown"
        grouped.setdefault(src, []).append(ts)

    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    # User requested: black background and black text.
    header_font = Font(name="Consolas", size=9, color="000000")
    value_font = Font(name="Consolas", size=9, color="000000")
    code_font = Font(name="Consolas", size=9, color="000000", bold=True)
    formula_font = Font(name="Consolas", size=8, color="000000", italic=True)
    date_font = Font(name="Consolas", size=9, color="000000")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # Generate dates: start from end_date + 2, descending to start_date
    date_top = end_dt + pd.DateOffset(days=2)
    dates = pd.date_range(start=start_dt, end=date_top, freq="D").sort_values(
        ascending=False
    )

    for sheet_source, ts_list in grouped.items():
        ws = wb.create_sheet(title=(sheet_source or "Unknown")[:31])

        # Row 1: Header info
        ws.cell(row=1, column=1, value=sheet_source).font = Font(
            name="Consolas", size=10, color="00BFFF", bold=True
        )
        ws.cell(row=1, column=1).fill = black_fill
        # C1 / D1 as real dates with date formatting
        ws.cell(row=1, column=3, value=start_dt.to_pydatetime()).font = value_font
        ws.cell(row=1, column=3).number_format = "yyyy-mm-dd"
        ws.cell(row=1, column=3).fill = black_fill
        ws.cell(row=1, column=4, value=end_dt.to_pydatetime()).font = value_font
        ws.cell(row=1, column=4).number_format = "yyyy-mm-dd"
        ws.cell(row=1, column=4).fill = black_fill

        # Metadata row labels (col A)
        row_labels = {
            2: "source_ticker",
            3: "source_field",
            4: "source_code",
            5: "asset_class",
            6: "source",
            7: "name",
            8: "code",
        }
        for r, label in row_labels.items():
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = header_font
            cell.fill = black_fill

        # Fill columns B+ with timeseries metadata
        for col_idx, ts in enumerate(ts_list, start=2):
            sc = str(ts.source_code) if ts.source_code else ":"
            parts = sc.rsplit(":", 1) if ":" in sc else [sc, ""]
            ticker = parts[0] if len(parts) > 0 else ""
            field = parts[1] if len(parts) > 1 else ""

            ws.cell(row=2, column=col_idx, value=ticker).font = value_font
            ws.cell(row=2, column=col_idx).fill = black_fill
            ws.cell(row=3, column=col_idx, value=field).font = value_font
            ws.cell(row=3, column=col_idx).fill = black_fill
            ws.cell(row=4, column=col_idx, value=sc).font = value_font
            ws.cell(row=4, column=col_idx).fill = black_fill
            ws.cell(row=5, column=col_idx, value=ts.asset_class or "").font = value_font
            ws.cell(row=5, column=col_idx).fill = black_fill
            ws.cell(row=6, column=col_idx, value=ts.source or "").font = value_font
            ws.cell(row=6, column=col_idx).fill = black_fill
            ws.cell(row=7, column=col_idx, value=ts.name or "").font = value_font
            ws.cell(row=7, column=col_idx).fill = black_fill
            ws.cell(row=8, column=col_idx, value=ts.code or "").font = code_font
            ws.cell(row=8, column=col_idx).fill = black_fill

            # Universal formula — column letter swapped via template
            from openpyxl.utils import get_column_letter

            col = get_column_letter(col_idx)
            formula_text = _DOWNLOAD_FORMULA_TEMPLATE.replace("__C__", col)

            # Keep as normal Excel formula (no array-brace rendering).
            ws.cell(row=9, column=col_idx, value=formula_text).font = formula_font
            ws.cell(row=9, column=col_idx).fill = black_fill

        # Date column (A9+) — descending from end_date + 2
        for i, d in enumerate(dates):
            ws.cell(row=9 + i, column=1, value=d.strftime("%Y-%m-%d")).font = date_font
            ws.cell(row=9 + i, column=1).fill = black_fill

        # Force black fill across the full used template area.
        max_row = 8 + len(dates)
        max_col = len(ts_list) + 1
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = black_fill

        # Column widths
        ws.column_dimensions["A"].width = 16
        for col_idx in range(2, len(ts_list) + 2):
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = 14

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"timeseries_template_{end_dt.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/timeseries/upload_template_data")
async def upload_template_data(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: SessionType = Depends(get_db),
    current_user: User = Depends(get_current_admin_user),
):
    """Upload Excel file with timeseries data in the background."""
    from ix.api.routers.task import start_process

    contents = await file.read()
    pid = start_process(
        f"Timeseries Upload: {file.filename}",
        user_id=str(getattr(current_user, "id", "") or ""),
    )

    background_tasks.add_task(_run_upload_background_task, pid, contents)

    return {"message": "Process started in background.", "task_id": pid}


def _run_upload_background_task(pid: str, contents: bytes):
    """Actual worker for timeseries data upload."""
    import io
    import math
    import pandas as pd
    from datetime import datetime, date
    from openpyxl import load_workbook
    from ix.db.models import Timeseries, TimeseriesData
    from ix.db.conn import Session
    from ix.api.routers.task import update_process, ProcessStatus

    logger.info(f"[Task:{pid}] Starting background timeseries upload.")
    try:
        with Session() as db:
            logger.info(f"[Task:{pid}] Parsing workbook contents...")
            update_process(pid, message="Parsing Excel file...")
            try:
                wb = load_workbook(io.BytesIO(contents), data_only=True)
            except Exception as e:
                logger.error(f"[Task:{pid}] Failed to load workbook: {e}")
                update_process(
                    pid,
                    status=ProcessStatus.FAILED,
                    message=f"Invalid Excel format: {str(e)}",
                )
                return

            # Fetch mappings and prepare
            ts_map = {
                ts.code: ts.id for ts in db.query(Timeseries.code, Timeseries.id).all()
            }
            logger.info(
                f"[Task:{pid}] Loaded mapping for {len(ts_map)} timeseries codes."
            )
            updated_count = 0
            total_points = 0
            sheets = wb.sheetnames

            for s_idx, sheet_name in enumerate(sheets):
                ws = wb[sheet_name]
                sheet_progress = f"{s_idx + 1}/{len(sheets)}"
                logger.info(
                    f"[Task:{pid}] Processing sheet: {sheet_name} ({sheet_progress})"
                )

                total_cols = ws.max_column - 1
                for col_idx in range(2, ws.max_column + 1):
                    col_num = col_idx - 1
                    code_cell = ws.cell(row=8, column=col_idx).value

                    if not code_cell:
                        continue

                    ts_code = str(code_cell).strip()
                    ts_id = ts_map.get(ts_code)

                    # Log every column to show activity
                    logger.info(
                        f"[Task:{pid}]   → Scanning '{ts_code}' (Col {col_num}/{total_cols})"
                    )
                    update_process(
                        pid,
                        message=f"Processing {sheet_name}: '{ts_code}' ({col_num}/{total_cols})",
                        progress=sheet_progress,
                    )

                    if not ts_id:
                        logger.warning(
                            f"[Task:{pid}]     ! Code '{ts_code}' not found in database. Skipping."
                        )
                        continue

                    # Fetch/Create data record
                    ts_data = (
                        db.query(TimeseriesData)
                        .filter(TimeseriesData.timeseries_id == ts_id)
                        .first()
                    )
                    if not ts_data:
                        ts_data = TimeseriesData(timeseries_id=ts_id, data={})
                        db.add(ts_data)
                        db.flush()

                    current_data = dict(ts_data.data) if ts_data.data else {}
                    has_changes = False
                    col_points = 0

                    # Iterate rows
                    for row_idx in range(9, ws.max_row + 1):
                        date_cell = ws.cell(row=row_idx, column=1).value
                        val_cell = ws.cell(row=row_idx, column=col_idx).value

                        if val_cell is None:
                            continue

                        date_str = None
                        if isinstance(date_cell, (datetime, date)):
                            date_str = date_cell.strftime("%Y-%m-%d")
                        elif isinstance(date_cell, str):
                            try:
                                date_str = pd.to_datetime(date_cell).strftime(
                                    "%Y-%m-%d"
                                )
                            except:
                                pass

                        if not date_str:
                            continue

                        try:
                            val = float(val_cell)
                            if math.isnan(val):
                                continue
                        except:
                            continue

                        current_data[date_str] = val
                        has_changes = True
                        col_points += 1

                    if has_changes:
                        logger.info(
                            f"[Task:{pid}]     ✓ Updated {ts_code}: +{col_points} points"
                        )
                        ts_data.data = current_data
                        if hasattr(ts_data, "updated"):
                            ts_data.updated = datetime.now()
                        updated_count += 1
                        total_points += col_points
                        db.flush()
                    else:
                        logger.info(f"[Task:{pid}]     · No new data for {ts_code}")

            db.commit()
            final_msg = f"Completed: {updated_count} series, {total_points} pts."
            logger.info(f"[Task:{pid}] {final_msg}")
            update_process(
                pid,
                status=ProcessStatus.COMPLETED,
                message=final_msg,
                progress=f"{len(sheets)}/{len(sheets)}",
            )

    except Exception as e:
        import traceback

        traceback.print_exc()
        logger.error(f"[Task:{pid}] Critical runtime error: {e}")
        update_process(
            pid, status=ProcessStatus.FAILED, message=f"Runtime error: {str(e)}"
        )


@router.post("/timeseries")
async def create_or_update_timeseries_bulk(
    payload: List[TimeseriesCreate],
    db: SessionType = Depends(get_db),
):
    """
    POST /api/timeseries - Create new or update existing timeseries metadata (bulk).

    Request body: List of timeseries objects. Each object can have:
    - id: Optional UUID string to update existing timeseries by ID
    - code: Required for new timeseries, optional if id is provided
    - Other fields: name, provider, asset_class, category, etc.

    Update logic:
    1. If 'id' is provided, find by ID first
    2. If not found by ID and 'code' is provided, find by code
    3. If not found, create new (requires 'code')
    """
    ensure_connection()

    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload provided.")

    updated_codes = []
    created_codes = []
    errors = []

    for ts_data in payload:
        try:
            ts = None

            # If ID is provided, try to find by ID first
            if ts_data.id:
                ts = db.query(Timeseries).filter(Timeseries.id == ts_data.id).first()
                if ts:
                    updated_codes.append(ts.code if ts.code else ts_data.id)

            # If not found by ID, try to find by code
            if ts is None and ts_data.code:
                ts = (
                    db.query(Timeseries).filter(Timeseries.code == ts_data.code).first()
                )
                if ts:
                    updated_codes.append(ts_data.code)

            # If still not found, create new (requires code)
            if ts is None:
                if not ts_data.code:
                    errors.append(
                        "Either 'id' or 'code' is required to create/update timeseries"
                    )
                    continue
                # Create new timeseries
                ts = Timeseries(code=ts_data.code)
                db.add(ts)
                db.flush()
                created_codes.append(ts_data.code)

            # Update fields
            # Update code if provided (only if different from current)
            if ts_data.code and ts_data.code != ts.code:
                ts.code = ts_data.code

            if ts_data.name is not None:
                ts.name = str(ts_data.name)[:200] if ts_data.name else None
            if ts_data.provider is not None:
                ts.provider = str(ts_data.provider)[:100] if ts_data.provider else None
            if ts_data.asset_class is not None:
                ts.asset_class = (
                    str(ts_data.asset_class)[:50] if ts_data.asset_class else None
                )
            if ts_data.category is not None:
                ts.category = str(ts_data.category)[:100] if ts_data.category else None
            if ts_data.source is not None:
                ts.source = str(ts_data.source)[:100] if ts_data.source else None
            if ts_data.source_code is not None:
                ts.source_code = (
                    str(ts_data.source_code)[:2000] if ts_data.source_code else None
                )
            if ts_data.frequency is not None:
                ts.frequency = (
                    str(ts_data.frequency)[:20] if ts_data.frequency else None
                )
            if ts_data.unit is not None:
                ts.unit = str(ts_data.unit)[:50] if ts_data.unit else None
            if ts_data.scale is not None:
                ts.scale = ts_data.scale
            if ts_data.currency is not None:
                ts.currency = str(ts_data.currency)[:10] if ts_data.currency else None
            if ts_data.country is not None:
                ts.country = str(ts_data.country)[:100] if ts_data.country else None
            if ts_data.remark is not None:
                ts.remark = str(ts_data.remark) if ts_data.remark else None
            if ts_data.favorite is not None:
                ts.favorite = ts_data.favorite

            db.commit()

        except Exception as e:
            identifier = ts_data.id or ts_data.code or "unknown"
            logger.error(f"Error updating timeseries {identifier}: {e}")
            errors.append(f"Error updating {identifier}: {str(e)}")
            db.rollback()
            continue

    response = {
        "message": f"Successfully processed {len(payload)} timeseries objects.",
        "created_codes": created_codes,
        "created_count": len(created_codes),
        "updated_codes": updated_codes,
        "updated_count": len(updated_codes),
    }

    if errors:
        response["errors"] = errors
        response["error_count"] = len(errors)

    return response


@router.put("/timeseries/{code}", response_model=TimeseriesResponse)
async def update_timeseries(
    code: str, payload: TimeseriesUpdate, db: SessionType = Depends(get_db)
):
    """
    PUT /api/timeseries/{code} - Update a single timeseries metadata entry.

    This endpoint follows REST best practices by using PUT for updates.
    Only provided fields are updated (no payload -> 400).
    """
    ensure_connection()

    ts = db.query(Timeseries).filter(Timeseries.code == code).first()
    if not ts:
        raise HTTPException(
            status_code=404, detail=f"Timeseries with code '{code}' not found"
        )

    update_fields = payload.model_dump(exclude_unset=True)
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields provided to update.")

    try:
        # Apply changes
        try:
            _apply_timeseries_updates(ts, update_fields)
        except (TypeError, ValueError):
            raise HTTPException(
                status_code=400, detail="Invalid field type in payload."
            )

        ts.updated = datetime.now()
        db.commit()
        db.refresh(ts)

        return TimeseriesResponse(
            id=str(ts.id),
            code=ts.code,
            name=ts.name,
            provider=ts.provider,
            asset_class=ts.asset_class,
            category=ts.category,
            start=ts.start,
            end=ts.end,
            num_data=ts.num_data,
            source=ts.source,
            source_code=getattr(ts, "source_code", None),
            frequency=ts.frequency,
            unit=getattr(ts, "unit", None),
            scale=getattr(ts, "scale", None),
            currency=getattr(ts, "currency", None),
            country=getattr(ts, "country", None),
            remark=getattr(ts, "remark", None),
            favorite=getattr(ts, "favorite", False),
        )
    except HTTPException:
        # Already constructed, just re-raise
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to update timeseries {code}: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to update timeseries '{code}'"
        )


@router.post("/timeseries/{code}", response_model=TimeseriesResponse, status_code=201)
async def create_timeseries(
    code: str, payload: TimeseriesCreate, db: SessionType = Depends(get_db)
):
    """
    POST /api/timeseries/{code} - Create a single timeseries metadata entry.

    - Returns 409 if the code already exists.
    - Body fields mirror the bulk create schema; path code overrides body code.
    """
    ensure_connection()

    # Validate existence
    existing = db.query(Timeseries).filter(Timeseries.code == code).first()
    if existing:
        raise HTTPException(
            status_code=409, detail=f"Timeseries with code '{code}' already exists"
        )

    # Build data from payload with path code taking precedence
    create_data = payload.model_dump(exclude_unset=True)
    create_data["code"] = code  # enforce path code

    try:
        ts = Timeseries(code=code)
        _apply_timeseries_updates(ts, create_data)
        ts.created = datetime.now()
        ts.updated = datetime.now()

        db.add(ts)
        db.commit()
        db.refresh(ts)

        return TimeseriesResponse(
            id=str(ts.id),
            code=ts.code,
            name=ts.name,
            provider=ts.provider,
            asset_class=ts.asset_class,
            category=ts.category,
            start=ts.start,
            end=ts.end,
            num_data=ts.num_data,
            source=ts.source,
            source_code=getattr(ts, "source_code", None),
            frequency=ts.frequency,
            unit=getattr(ts, "unit", None),
            scale=getattr(ts, "scale", None),
            currency=getattr(ts, "currency", None),
            country=getattr(ts, "country", None),
            remark=getattr(ts, "remark", None),
            favorite=getattr(ts, "favorite", False),
        )
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to create timeseries {code}: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to create timeseries '{code}'"
        )


@router.delete("/timeseries/{code}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_timeseries(
    code: str,
    db: SessionType = Depends(get_db),
    current_user: User = Depends(get_current_admin_user),
):
    """
    DELETE /api/timeseries/{code} - Delete a timeseries by code.

    Admin-only endpoint.
    """
    ensure_connection()

    ts = db.query(Timeseries).filter(Timeseries.code == code).first()
    if not ts:
        raise HTTPException(
            status_code=404, detail=f"Timeseries with code '{code}' not found"
        )

    try:
        db.delete(ts)
        db.commit()
        logger.info(f"Admin {current_user.email} deleted timeseries: {code}")
        return
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to delete timeseries {code}: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to delete timeseries '{code}'"
        )


@router.get("/timeseries/{timeseries_id}", response_model=TimeseriesResponse)
async def get_timeseries_by_id(
    timeseries_id: str,
    db: SessionType = Depends(get_db),
):
    """
    GET /api/timeseries/{id} - Get detailed timeseries information by ID (code) with full data.
    """
    ensure_connection()

    ts = db.query(Timeseries).filter(Timeseries.code == timeseries_id).first()

    if not ts:
        raise HTTPException(status_code=404, detail="Timeseries not found")

    return TimeseriesResponse(
        id=str(ts.id),
        code=ts.code,
        name=ts.name,
        provider=ts.provider,
        asset_class=ts.asset_class,
        category=ts.category,
        start=ts.start,
        end=ts.end,
        num_data=ts.num_data,
        source=ts.source,
        source_code=getattr(ts, "source_code", None),
        frequency=ts.frequency,
        unit=getattr(ts, "unit", None),
        scale=getattr(ts, "scale", None),
        currency=getattr(ts, "currency", None),
        country=getattr(ts, "country", None),
        remark=getattr(ts, "remark", None),
        favorite=getattr(ts, "favorite", False),
    )


@router.get("/timeseries/code/{code}", response_model=TimeseriesResponse)
async def get_timeseries_by_code(
    code: str,
    db: SessionType = Depends(get_db),
):
    """
    GET /api/timeseries/code/{code} - Get timeseries by its code.
    """
    ensure_connection()

    ts = db.query(Timeseries).filter(Timeseries.code == code).first()

    if not ts:
        raise HTTPException(
            status_code=404, detail=f"Timeseries with code '{code}' not found"
        )

    return TimeseriesResponse(
        id=str(ts.id),
        code=ts.code,
        name=ts.name,
        provider=ts.provider,
        asset_class=ts.asset_class,
        category=ts.category,
        start=ts.start,
        end=ts.end,
        num_data=ts.num_data,
        source=ts.source,
        source_code=getattr(ts, "source_code", None),
        frequency=ts.frequency,
        unit=getattr(ts, "unit", None),
        scale=getattr(ts, "scale", None),
        currency=getattr(ts, "currency", None),
        country=getattr(ts, "country", None),
        remark=getattr(ts, "remark", None),
        favorite=getattr(ts, "favorite", False),
    )


@router.get("/timeseries/favorites")
async def get_favorite_timeseries_data(
    start_date: Optional[str] = Query(None, description="Start date (ISO-8601)"),
    end_date: Optional[str] = Query(None, description="End date (ISO-8601)"),
    db: SessionType = Depends(get_db),
):
    """
    GET /api/timeseries/favorites - Get all favorite timeseries data as a concatenated DataFrame.

    Returns column-oriented JSON format with dates sorted descending.
    """
    ensure_connection()

    try:
        # Query all favorite timeseries with data_record eagerly loaded
        favorite_timeseries = (
            db.query(Timeseries)
            .options(joinedload(Timeseries.data_record))
            .filter(Timeseries.favorite == True)
            .all()
        )

        if not favorite_timeseries:
            return Response(
                content=json.dumps({"Date": []}, ensure_ascii=False),
                media_type="application/json",
            )

        # Collect all series data as pandas Series for concatenation
        series_list = []

        for ts in favorite_timeseries:
            try:
                ts_code = ts.code
                data_record = ts._get_or_create_data_record(db)
                column_data = (
                    data_record.data if data_record and data_record.data else {}
                )
                frequency = ts.frequency

                # Convert JSONB dict to pandas Series
                if column_data and isinstance(column_data, dict):
                    ts_data = pd.Series(column_data)
                    try:
                        ts_data.index = pd.to_datetime(ts_data.index)
                        ts_data = pd.to_numeric(ts_data, errors="coerce").dropna()
                        ts_data.name = ts_code
                        if frequency and len(ts_data) > 0:
                            ts_data = (
                                ts_data.sort_index()
                                .resample(str(frequency))
                                .last()
                                .dropna()
                            )
                        else:
                            ts_data = ts_data.sort_index()

                        if not ts_data.empty:
                            series_list.append(ts_data)
                    except Exception:
                        try:
                            valid_dates = pd.to_datetime(ts_data.index, errors="coerce")
                            ts_data = ts_data[valid_dates.notna()]
                            ts_data.index = pd.to_datetime(ts_data.index)
                            ts_data = pd.to_numeric(ts_data, errors="coerce").dropna()
                            ts_data.name = ts_code
                            ts_data = ts_data.sort_index()
                            if not ts_data.empty:
                                series_list.append(ts_data)
                        except Exception:
                            logger.warning(
                                f"Error processing timeseries {ts_code}: could not convert to Series"
                            )
                            continue
            except Exception as e:
                logger.warning(f"Error processing favorite timeseries {ts.code}: {e}")
                continue

        # Concatenate all series into a DataFrame
        if series_list:
            df = pd.concat(series_list, axis=1)
            df.index.name = "Date"

            # Optional date slicing
            start_ts = (
                pd.to_datetime(start_date, errors="coerce") if start_date else None
            )
            end_ts = pd.to_datetime(end_date, errors="coerce") if end_date else None

            # Normalize timezone info
            try:
                df.index = pd.DatetimeIndex(df.index).tz_localize(None)
            except Exception:
                try:
                    df.index = pd.DatetimeIndex(df.index).tz_convert(None)
                except Exception:
                    pass

            df = df.resample("D").last()

            if isinstance(start_ts, pd.Timestamp):
                try:
                    start_ts = start_ts.tz_localize(None)
                except Exception:
                    try:
                        start_ts = start_ts.tz_convert(None)
                    except Exception:
                        pass

            if isinstance(end_ts, pd.Timestamp):
                try:
                    end_ts = end_ts.tz_localize(None)
                except Exception:
                    try:
                        end_ts = end_ts.tz_convert(None)
                    except Exception:
                        pass

            # Apply slicing if bounds are valid
            if isinstance(start_ts, pd.Timestamp):
                df = df[df.index >= start_ts]
            if isinstance(end_ts, pd.Timestamp):
                df = df[df.index <= end_ts]

            # Sort dates descending
            df = df.sort_index(ascending=True)

            # Convert to column-oriented format
            df_indexed = df.reset_index()
            column_dict = OrderedDict()

            for col in df_indexed.columns:
                values = df_indexed[col].tolist()
                cleaned_values = []
                for v in values:
                    if v is None or (isinstance(v, float) and math.isnan(v)):
                        cleaned_values.append(None)
                    elif isinstance(v, (pd.Timestamp, datetime)):
                        cleaned_values.append(v.isoformat())
                    else:
                        cleaned_values.append(v)
                column_dict[col] = cleaned_values

            return Response(
                content=json.dumps(column_dict, ensure_ascii=False),
                media_type="application/json",
            )
        else:
            return Response(
                content=json.dumps({"Date": []}, ensure_ascii=False),
                media_type="application/json",
            )

    except Exception as e:
        logger.exception(f"Error retrieving favorite timeseries: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


async def _parse_codes_from_request(request: Request) -> List[str]:
    """Parse codes from request body, query parameters, or X-Codes header."""
    codes = []

    # Try to get from request body (works for POST, and GET with body in some clients)
    try:
        body = await request.json()
        if isinstance(body, list):
            codes = [str(c).strip() for c in body if str(c).strip()]
        elif isinstance(body, dict) and body:
            raw_codes = body.get("codes")
            if isinstance(raw_codes, list):
                codes = [str(c).strip() for c in raw_codes if str(c).strip()]
            elif isinstance(raw_codes, str):
                codes = [c.strip() for c in raw_codes.split(",") if c.strip()]
    except Exception:
        pass

    # Fallback to query parameters (for GET requests)
    if not codes:
        query_codes = request.query_params.get("codes")
        if query_codes:
            codes = [c.strip() for c in query_codes.split(",") if c.strip()]

    # Fallback to header if body and query didn't work
    if not codes:
        header_codes = request.headers.get("X-Codes")
        if header_codes:
            # Try to parse as JSON first (handles codes with commas and special chars)
            try:
                parsed = json.loads(header_codes)
                if isinstance(parsed, list):
                    codes = [str(c).strip() for c in parsed if str(c).strip()]
                    logger.debug(
                        f"Parsed {len(codes)} codes from JSON array in X-Codes header"
                    )
                elif isinstance(parsed, dict) and "codes" in parsed:
                    codes_list = parsed["codes"]
                    if isinstance(codes_list, list):
                        codes = [str(c).strip() for c in codes_list if str(c).strip()]
                        logger.debug(
                            f"Parsed {len(codes)} codes from JSON dict in X-Codes header"
                        )
                    elif isinstance(codes_list, str):
                        codes = [c.strip() for c in codes_list.split(",") if c.strip()]
                        logger.warning(
                            "X-Codes header contains dict with string 'codes' - using comma-split (may break)"
                        )
                else:
                    logger.warning(
                        f"X-Codes JSON parsed but unexpected format: {type(parsed)}"
                    )
            except (json.JSONDecodeError, ValueError) as e:
                # Check if header looks like it might contain complex expressions with commas
                # If it contains function calls, quotes, or parentheses, it's likely JSON should be used
                has_complex_expressions = any(
                    char in header_codes
                    for char in [
                        "(",
                        ")",
                        "{",
                        "}",
                        '"',
                        "'",
                        "Series(",
                        "MultiSeries(",
                    ]
                )

                if has_complex_expressions:
                    logger.error(
                        f"X-Codes header contains complex expressions but is not valid JSON. "
                        f'Please send codes as JSON array: ["code1", "code2"]. '
                        f"Error: {e}. Header preview: {header_codes[:200]}..."
                    )
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "X-Codes header must be a JSON array when codes contain commas or special characters. "
                            'Example: X-Codes: ["CODE1", "Series(\'CODE2\').pct_change()"]'
                        ),
                    )

                # Fallback to comma-separated for simple codes only
                logger.warning(
                    f"X-Codes header is not valid JSON (falling back to comma-separated): {e}. "
                    f"Header value starts with: {header_codes[:100]}..."
                )
                codes = [c.strip() for c in header_codes.split(",") if c.strip()]

    if not codes:
        raise HTTPException(
            status_code=400,
            detail="No codes provided. Provide JSON body with 'codes' list, 'codes' query parameter, or 'X-Codes' header.",
        )

    # Filter out reserved column names
    RESERVED_NAMES = {"Date", "date"}  # Reserved for index column
    codes = [c for c in codes if c not in RESERVED_NAMES]

    if not codes:
        raise HTTPException(
            status_code=400,
            detail="No valid codes provided after filtering reserved names.",
        )

    # Deduplicate while preserving order
    seen_codes = set()
    unique_codes = []
    for code in codes:
        if code not in seen_codes:
            unique_codes.append(code)
            seen_codes.add(code)

    return unique_codes


def _process_database_timeseries(
    ts: Timeseries, db: SessionType
) -> Optional[pd.Series]:
    """Process a timeseries from the database and return as pandas Series."""
    try:
        data_record = ts._get_or_create_data_record(db)
        column_data = data_record.data if data_record and data_record.data else {}
        frequency = ts.frequency

        if column_data and isinstance(column_data, dict):
            ts_data = pd.Series(column_data)
            try:
                ts_data.index = pd.to_datetime(ts_data.index)
                ts_data = pd.to_numeric(ts_data, errors="coerce").dropna()
                ts_data.name = ts.code
                if frequency and len(ts_data) > 0:
                    ts_data = (
                        ts_data.sort_index().resample(str(frequency)).last().dropna()
                    )
                else:
                    ts_data = ts_data.sort_index()

                if not ts_data.empty:
                    return ts_data
            except Exception:
                try:
                    valid_dates = pd.to_datetime(ts_data.index, errors="coerce")
                    ts_data = ts_data[valid_dates.notna()]
                    ts_data.index = pd.to_datetime(ts_data.index)
                    ts_data = pd.to_numeric(ts_data, errors="coerce").dropna()
                    ts_data.name = ts.code
                    ts_data = ts_data.sort_index()
                    if not ts_data.empty:
                        return ts_data
                except Exception:
                    logger.warning(
                        f"Error processing timeseries {ts.code}: could not convert to Series"
                    )
                    return None
    except Exception as e:
        logger.warning(f"Error processing database timeseries {ts.code}: {e}")
        return None

    return None


def _evaluate_expression(
    code: str, start_date: Optional[str], end_date: Optional[str]
) -> List[pd.Series]:
    """Evaluate code as Python expression and return list of Series."""
    try:
        logger.info(
            f"Code {code} not found in database, attempting to evaluate as expression"
        )
        evaluated_series = eval(code)
        series_list = []

        # Handle both Series and DataFrame results
        if isinstance(evaluated_series, pd.Series):
            evaluated_series.name = code
            if not evaluated_series.empty:
                # Apply date filtering if provided
                if start_date:
                    start_dt = pd.to_datetime(start_date, errors="coerce")
                    if start_dt:
                        evaluated_series = evaluated_series.loc[
                            evaluated_series.index >= start_dt
                        ]
                if end_date:
                    end_dt = pd.to_datetime(end_date, errors="coerce")
                    if end_dt:
                        evaluated_series = evaluated_series.loc[
                            evaluated_series.index <= end_dt
                        ]
                series_list.append(evaluated_series)
        elif isinstance(evaluated_series, pd.DataFrame):
            # If DataFrame, convert each column to a series
            for col in evaluated_series.columns:
                col_series = evaluated_series[col].copy()
                col_series.name = col
                # Apply date filtering if provided
                if start_date:
                    start_dt = pd.to_datetime(start_date, errors="coerce")
                    if start_dt:
                        col_series = col_series.loc[col_series.index >= start_dt]
                if end_date:
                    end_dt = pd.to_datetime(end_date, errors="coerce")
                    if end_dt:
                        col_series = col_series.loc[col_series.index <= end_dt]
                if not col_series.empty:
                    series_list.append(col_series)
        else:
            logger.warning(
                f"Evaluated expression {code} did not return a Series or DataFrame"
            )

        return series_list
    except Exception as e:
        logger.warning(
            f"Code {code} not found in database and failed to evaluate as expression: {e}"
        )
        return []


def _normalize_timezone(ts: pd.Timestamp) -> pd.Timestamp:
    """Normalize timezone info from a Timestamp."""
    try:
        return ts.tz_localize(None)
    except Exception:
        try:
            return ts.tz_convert(None)
        except Exception:
            return ts


def _format_dataframe_response(
    df: pd.DataFrame,
    series_list: List[pd.Series],
    start_date: Optional[str],
    end_date: Optional[str],
) -> Response:
    """Format DataFrame into column-oriented JSON response."""
    df.index.name = "Date"

    # Optional date bounds
    start_ts = pd.to_datetime(start_date, errors="coerce") if start_date else None
    end_ts = pd.to_datetime(end_date, errors="coerce") if end_date else None

    # Normalize timezone info
    try:
        df.index = pd.DatetimeIndex(df.index).tz_localize(None)
    except Exception:
        try:
            df.index = pd.DatetimeIndex(df.index).tz_convert(None)
        except Exception:
            pass

    # Resample to daily frequency
    df = df.resample("D").last()

    if isinstance(start_ts, pd.Timestamp):
        start_ts = _normalize_timezone(start_ts)
    if isinstance(end_ts, pd.Timestamp):
        end_ts = _normalize_timezone(end_ts)

    # Apply slicing if bounds are valid
    if isinstance(start_ts, pd.Timestamp):
        df = df[df.index >= start_ts]
    if isinstance(end_ts, pd.Timestamp):
        df = df[df.index <= end_ts]

    # Reorder columns to preserve input order
    present_in_order = [s.name for s in series_list if s.name in df.columns]
    seen_present = set()
    present_in_order = [
        c for c in present_in_order if not (c in seen_present or seen_present.add(c))
    ]
    df = df[present_in_order]

    # Sort dates ascending as requested
    df = df.sort_index(ascending=True)

    # Convert to column-oriented format
    df_indexed = df.reset_index()
    column_dict = OrderedDict()

    for col in df_indexed.columns:
        values = df_indexed[col].tolist()
        cleaned_values = []
        for v in values:
            if v is None or (isinstance(v, float) and math.isnan(v)):
                cleaned_values.append(None)
            elif isinstance(v, (pd.Timestamp, datetime)):
                cleaned_values.append(v.isoformat())
            else:
                cleaned_values.append(v)
        column_dict[col] = cleaned_values

    return Response(
        content=json.dumps(column_dict, ensure_ascii=False),
        media_type="application/json",
    )


@router.get("/timeseries.custom")
@router.post("/timeseries.custom")
async def get_custom_timeseries_data(
    request: Request,
    start_date: Optional[str] = Query(None, description="Start date (ISO-8601)"),
    end_date: Optional[str] = Query(None, description="End date (ISO-8601)"),
    db: SessionType = Depends(get_db),
):
    """
    GET/POST /api/timeseries.custom - Get selected timeseries data based on provided codes.

    Accepts JSON body (not in URL):
    { "codes": ["CODE1", "CODE2", ...] }
    or
    { "codes": "CODE1, CODE2, ..." }

    Also accepts X-Codes header:
    - JSON format (recommended for codes with commas): X-Codes: ["CODE1", "CODE2, WITH COMMA", "CODE3"]
    - Comma-separated (legacy, breaks if codes contain commas): X-Codes: CODE1, CODE2, CODE3

    Behavior:
    - Queries each code one by one from the Timeseries table
    - If code is found in database, uses stored data
    - If code is NOT found, attempts to evaluate it as a Python expression (e.g., Series('CODE'))
    - Supports mixing database timeseries with dynamically computed series

    Examples:
    - ["AAPL", "MSFT"] - database codes
    - ["AAPL", "Series('SPY')"] - mix of database and expression
    - ["Series('AAPL').pct_change()"] - pure expression

    Note: For codes containing commas or special characters, use JSON body or JSON in X-Codes header.

    Returns column-oriented format with dates sorted descending.
    """
    ensure_connection()

    try:
        codes = await _parse_codes_from_request(request)
        logger.debug(f"Processing {len(codes)} codes: {codes}")
        series_list = []

        # Process each code one by one
        for code in codes:
            try:
                # First, try to find in Timeseries table
                ts = (
                    db.query(Timeseries)
                    .options(joinedload(Timeseries.data_record))
                    .filter(Timeseries.code == code)
                    .first()
                )

                if ts:
                    # Found in database - use database data
                    ts_data = _process_database_timeseries(ts, db)
                    if ts_data is not None:
                        series_list.append(ts_data)
                else:
                    # Not found in database - try to evaluate as expression
                    evaluated_series = _evaluate_expression(code, start_date, end_date)
                    series_list.extend(evaluated_series)
            except Exception as e:
                logger.warning(f"Error processing custom timeseries {code}: {e}")
                continue

        # Concatenate all series into a DataFrame
        if series_list:
            df = pd.concat(series_list, axis=1)
            return _format_dataframe_response(df, series_list, start_date, end_date)
        else:
            return Response(
                content=json.dumps({"Date": []}, ensure_ascii=False),
                media_type="application/json",
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error retrieving custom timeseries: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.post("/upload_data")
async def upload_data(
    payload: TimeseriesDataUpload,
    db: SessionType = Depends(get_db),
):
    """
    POST /api/upload_data - Upload timeseries data in bulk (updates existing data).
    """
    try:
        ensure_connection()

        # Convert to DataFrame
        df = pd.DataFrame(
            [{"date": d.date, "code": d.code, "value": d.value} for d in payload.data]
        )

        if df.empty:
            raise HTTPException(status_code=400, detail="No data provided")

        # Clean the data
        df["value"] = pd.to_numeric(df["value"], errors="coerce")
        df = df.dropna(subset=["value"])
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date", "code"])

        if df.empty:
            raise HTTPException(
                status_code=400, detail="No valid records to upload after cleaning."
            )

        # Pivot the DataFrame
        pivoted = (
            df.pivot(index="date", columns="code", values="value")
            .sort_index()
            .dropna(how="all", axis=1)
            .dropna(how="all", axis=0)
        )
        pivoted.index = pd.to_datetime(pivoted.index)

        # Update each timeseries in the database
        updated_codes = []
        not_found_codes = []
        for code in pivoted.columns:
            ts = db.query(Timeseries).filter(Timeseries.code == code).first()
            if ts is None:
                not_found_codes.append(code)
                continue

            # Get existing data to merge with new data
            data_record = ts._get_or_create_data_record(db)
            column_data = data_record.data if data_record and data_record.data else {}
            if column_data and len(column_data) > 0:
                data_dict = column_data if isinstance(column_data, dict) else {}
                existing_data = pd.Series(data_dict)
                if not existing_data.empty:
                    existing_data.index = pd.to_datetime(
                        existing_data.index, errors="coerce"
                    )
                    existing_data = existing_data.dropna()
            else:
                existing_data = pd.Series(dtype=float)

            # Store new data as a Series
            new_series = pivoted[code].dropna()

            if not existing_data.empty:
                combined_data = pd.concat([existing_data, new_series], axis=0)
                combined_data = combined_data[
                    ~combined_data.index.duplicated(keep="last")
                ]
                combined_data = combined_data.sort_index()
            else:
                combined_data = new_series

            # Convert combined_data to dict format with date strings for JSONB storage
            data_dict = {}
            for k, v in combined_data.to_dict().items():
                if hasattr(k, "date"):
                    date_str = str(k.date())
                elif isinstance(k, str):
                    date_str = k
                else:
                    date_str = str(pd.to_datetime(k).date())
                data_dict[date_str] = (
                    float(v)
                    if v is not None and not (isinstance(v, float) and pd.isna(v))
                    else None
                )

            # Update the timeseries data
            data_record.data = data_dict
            data_record.updated = datetime.now()
            ts.start = (
                combined_data.index.min().date()
                if len(combined_data.index) > 0
                else None
            )
            ts.end = (
                combined_data.index.max().date()
                if len(combined_data.index) > 0
                else None
            )
            ts.num_data = len(combined_data)
            ts.updated = datetime.now()

            db.commit()
            logger.info(
                f"Updated {code}: merged {len(new_series)} new points with {len(existing_data) if not existing_data.empty else 0} existing points"
            )
            updated_codes.append(code)

        response = {
            "message": f"Successfully received {len(payload.data)} records.",
            "updated_codes": updated_codes,
            "records_processed": len(df),
        }

        if not_found_codes:
            response["warning"] = f"Codes not found in database: {not_found_codes}"
            logger.warning(f"Codes not found: {not_found_codes}")

        return response

    except Exception as e:
        logger.exception("Failed to process data upload")
        raise HTTPException(
            status_code=500, detail=f"Failed to process data upload: {str(e)}"
        )


@router.post("/upload_data_columnar")
async def upload_data_columnar(
    payload: TimeseriesColumnarUpload,
    db: SessionType = Depends(get_db),
):
    """
    POST /api/upload_data_columnar - Upload timeseries data in efficient columnar format.

    This endpoint is optimized for large dataset uploads. Instead of sending
    individual {date, code, value} objects, send data in columnar format:

    {
        "dates": ["2024-01-01", "2024-01-02", ...],
        "columns": {
            "CODE1": [1.23, 4.56, null, ...],
            "CODE2": [7.89, 0.12, null, ...]
        }
    }

    This reduces payload size by 80-90% compared to the row-by-row format.
    """
    try:
        ensure_connection()

        # Validate input
        if not payload.dates or not payload.columns:
            raise HTTPException(status_code=400, detail="No data provided")

        # Direct DataFrame construction from columnar format (no pivot needed!)
        try:
            dates_index = pd.to_datetime(payload.dates, errors="coerce")
            df = pd.DataFrame(payload.columns, index=dates_index)
        except Exception as e:
            raise HTTPException(
                status_code=400, detail=f"Invalid data format: {str(e)}"
            )

        # Drop rows where all values are null and columns where all values are null
        df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)

        if df.empty:
            raise HTTPException(
                status_code=400, detail="No valid records to upload after cleaning."
            )

        # Track results
        updated_codes = []
        not_found_codes = []
        total_points_merged = 0

        # Process each column (timeseries code)
        for code in df.columns:
            ts = db.query(Timeseries).filter(Timeseries.code == code).first()
            if ts is None:
                not_found_codes.append(code)
                continue

            # Get existing data
            data_record = ts._get_or_create_data_record(db)
            column_data = data_record.data if data_record and data_record.data else {}

            if column_data and isinstance(column_data, dict):
                existing_data = pd.Series(column_data)
                if not existing_data.empty:
                    existing_data.index = pd.to_datetime(
                        existing_data.index, errors="coerce"
                    )
                    existing_data = existing_data.dropna()
            else:
                existing_data = pd.Series(dtype=float)

            # Get new data for this code
            new_series = df[code].dropna()

            # Merge new with existing (new values override)
            if not existing_data.empty:
                combined_data = pd.concat([existing_data, new_series], axis=0)
                combined_data = combined_data[
                    ~combined_data.index.duplicated(keep="last")
                ]
                combined_data = combined_data.sort_index()
            else:
                combined_data = new_series.sort_index()

            # Convert to JSONB-friendly dict format
            data_dict = {}
            for k, v in combined_data.to_dict().items():
                if hasattr(k, "date"):
                    date_str = str(k.date())
                elif isinstance(k, str):
                    date_str = k
                else:
                    date_str = str(pd.to_datetime(k).date())
                data_dict[date_str] = (
                    float(v)
                    if v is not None and not (isinstance(v, float) and pd.isna(v))
                    else None
                )

            # Update timeseries record
            data_record.data = data_dict
            data_record.updated = datetime.now()
            ts.start = (
                combined_data.index.min().date()
                if len(combined_data.index) > 0
                else None
            )
            ts.end = (
                combined_data.index.max().date()
                if len(combined_data.index) > 0
                else None
            )
            ts.num_data = len(combined_data)
            ts.updated = datetime.now()

            total_points_merged += len(new_series)
            updated_codes.append(code)

        # Single commit for all codes (much faster than per-code commits)
        db.commit()

        logger.info(
            f"Columnar upload: updated {len(updated_codes)} codes with {total_points_merged} total data points"
        )

        response = {
            "message": f"Successfully processed {len(payload.dates)} dates × {len(payload.columns)} columns.",
            "updated_codes": updated_codes,
            "total_points_merged": total_points_merged,
        }

        if not_found_codes:
            response["warning"] = f"Codes not found in database: {not_found_codes}"
            logger.warning(f"Codes not found: {not_found_codes}")

        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to process columnar data upload")
        raise HTTPException(
            status_code=500, detail=f"Failed to process data upload: {str(e)}"
        )
