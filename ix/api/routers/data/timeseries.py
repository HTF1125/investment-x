"""
Timeseries router for timeseries data management.
"""

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    Request,
    status,
    UploadFile,
    File,
)
from fastapi.responses import Response, StreamingResponse, FileResponse
from typing import Optional, List, Dict, Any
from collections import OrderedDict
import json
import math
import re
import pandas as pd
from datetime import datetime, date

from ix.api.schemas import (
    TimeseriesResponse,
    TimeseriesCreate,
    TimeseriesBulkUpdate,
    TimeseriesDataUpload,
    TimeseriesColumnarUpload,
    TimeseriesUpdate,
)
from ix.api.dependencies import get_db, get_current_admin_user, get_current_user
from ix.db.models import Timeseries
from ix.db.conn import ensure_connection, Session
from ix.db.models.user import User
from sqlalchemy.orm import joinedload, Session as SessionType
from ix.misc import get_logger
from ix.api.rate_limit import limiter as _limiter
from ix.utils.safe_expression import (
    TIMESERIES_EXPRESSION_CONTEXT,
    UnsafeExpressionError,
    safe_eval_expression,
    safe_exec_code,
)

logger = get_logger(__name__)

router = APIRouter()


def _apply_timeseries_updates(ts: Timeseries, data: dict) -> None:
    """Apply whitelisted updates to a Timeseries instance."""
    if "code" in data and data["code"] is not None:
        ts.code = str(data["code"])

    if "name" in data:
        name = data["name"]
        ts.name = str(name)[:200] if name else None
    if "provider" in data:
        provider = data["provider"]
        ts.provider = str(provider)[:100] if provider else None
    if "asset_class" in data:
        asset_class = data["asset_class"]
        ts.asset_class = str(asset_class)[:50] if asset_class else None
    if "category" in data:
        category = data["category"]
        ts.category = str(category)[:100] if category else None
    if "source" in data:
        source = data["source"]
        ts.source = str(source)[:100] if source else None
    if "source_code" in data:
        source_code = data["source_code"]
        ts.source_code = str(source_code)[:2000] if source_code else None
    if "frequency" in data:
        frequency = data["frequency"]
        ts.frequency = str(frequency)[:20] if frequency else None
    if "unit" in data:
        unit = data["unit"]
        ts.unit = str(unit)[:50] if unit else None
    if "scale" in data:
        scale = data["scale"]
        if scale is not None:
            ts.scale = int(scale)
        else:
            ts.scale = None
    if "currency" in data:
        currency = data["currency"]
        ts.currency = str(currency)[:10] if currency else None
    if "country" in data:
        country = data["country"]
        ts.country = str(country)[:100] if country else None
    if "remark" in data:
        remark = data["remark"]
        ts.remark = str(remark) if remark else None
    if "favorite" in data:
        ts.favorite = bool(data["favorite"]) if data["favorite"] is not None else None


@router.get("/timeseries", response_model=List[TimeseriesResponse])
def get_timeseries(
    limit: Optional[int] = Query(None, ge=1, le=1000, description="Limit number of results"),
    offset: int = Query(0, ge=0, description="Offset for pagination"),
    search: Optional[str] = Query(
        None,
        max_length=200,
        description="Search by code/name/source/category/provider/asset class/country",
    ),
    category: Optional[str] = Query(None, max_length=100, description="Filter by category"),
    asset_class: Optional[str] = Query(None, max_length=100, description="Filter by asset class"),
    provider: Optional[str] = Query(None, max_length=100, description="Filter by provider"),
    db: SessionType = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """
    GET /api/timeseries - List all timeseries with optional filtering and pagination.
    """
    ensure_connection()
    from sqlalchemy import and_, or_, case, func

    timeseries_query = db.query(Timeseries)

    # DB-native search for PostgreSQL, with fallback ranking for other dialects.
    if search:
        term = search.strip().lower()
        tokens = [t for t in re.split(r"\s+", term) if t]

        if tokens:
            code_l = func.lower(func.coalesce(Timeseries.code, ""))
            name_l = func.lower(func.coalesce(Timeseries.name, ""))
            source_l = func.lower(func.coalesce(Timeseries.source, ""))
            category_l = func.lower(func.coalesce(Timeseries.category, ""))
            provider_l = func.lower(func.coalesce(Timeseries.provider, ""))
            asset_class_l = func.lower(func.coalesce(Timeseries.asset_class, ""))
            country_l = func.lower(func.coalesce(Timeseries.country, ""))
            source_code_l = func.lower(func.coalesce(Timeseries.source_code, ""))

            dialect_name = (
                getattr(getattr(db, "bind", None), "dialect", None).name
                if getattr(getattr(db, "bind", None), "dialect", None) is not None
                else ""
            )

            if dialect_name == "postgresql":
                # Weighted FTS vector:
                # code/name strongest; metadata moderate; source_code lower.
                search_vector = (
                    func.setweight(
                        func.to_tsvector("simple", func.coalesce(Timeseries.code, "")),
                        "A",
                    )
                    .op("||")(
                        func.setweight(
                            func.to_tsvector(
                                "simple", func.coalesce(Timeseries.name, "")
                            ),
                            "A",
                        )
                    )
                    .op("||")(
                        func.setweight(
                            func.to_tsvector(
                                "simple", func.coalesce(Timeseries.source, "")
                            ),
                            "B",
                        )
                    )
                    .op("||")(
                        func.setweight(
                            func.to_tsvector(
                                "simple", func.coalesce(Timeseries.category, "")
                            ),
                            "B",
                        )
                    )
                    .op("||")(
                        func.setweight(
                            func.to_tsvector(
                                "simple", func.coalesce(Timeseries.provider, "")
                            ),
                            "C",
                        )
                    )
                    .op("||")(
                        func.setweight(
                            func.to_tsvector(
                                "simple", func.coalesce(Timeseries.asset_class, "")
                            ),
                            "C",
                        )
                    )
                    .op("||")(
                        func.setweight(
                            func.to_tsvector(
                                "simple", func.coalesce(Timeseries.country, "")
                            ),
                            "D",
                        )
                    )
                    .op("||")(
                        func.setweight(
                            func.to_tsvector(
                                "simple", func.coalesce(Timeseries.source_code, "")
                            ),
                            "D",
                        )
                    )
                )
                ts_query = func.plainto_tsquery("simple", term)

                # Keep prefix path so incremental typing ("sp", "usd") still feels responsive.
                prefix_match = or_(
                    code_l.like(f"{term}%"),
                    name_l.like(f"{term}%"),
                    source_l.like(f"{term}%"),
                    category_l.like(f"{term}%"),
                )
                if len(term) >= 3:
                    prefix_match = or_(prefix_match, source_code_l.like(f"{term}%"))

                fts_match = search_vector.op("@@")(ts_query)
                timeseries_query = timeseries_query.filter(or_(fts_match, prefix_match))

                rank_expr = (
                    func.ts_rank_cd(search_vector, ts_query)
                    + case((code_l == term, 5.0), else_=0.0)
                    + case((code_l.like(f"{term}%"), 2.6), else_=0.0)
                    + case((name_l.like(f"{term}%"), 1.2), else_=0.0)
                    + case((source_l.like(f"{term}%"), 0.6), else_=0.0)
                    + case((category_l.like(f"{term}%"), 0.5), else_=0.0)
                )

                timeseries_query = timeseries_query.order_by(
                    rank_expr.desc(),
                    Timeseries.favorite.desc(),
                    Timeseries.code.asc(),
                )
            else:
                # Fallback: token-aware LIKE ranking for non-PostgreSQL backends.
                token_filters = []
                for tok in tokens:
                    tok_like = f"%{tok}%"
                    token_columns = [
                        code_l.like(tok_like),
                        name_l.like(tok_like),
                        source_l.like(tok_like),
                        category_l.like(tok_like),
                        provider_l.like(tok_like),
                        asset_class_l.like(tok_like),
                        country_l.like(tok_like),
                    ]
                    if len(tok) >= 3:
                        token_columns.append(source_code_l.like(tok_like))
                    token_filters.append(or_(*token_columns))

                timeseries_query = timeseries_query.filter(and_(*token_filters))

                rank_expr = (
                    case((code_l == term, 1000), else_=0)
                    + case((code_l.like(f"{term}%"), 600), else_=0)
                    + case((name_l == term, 350), else_=0)
                    + case((name_l.like(f"{term}%"), 220), else_=0)
                    + case((source_code_l.like(f"{term}%"), 180), else_=0)
                    + case((code_l.like(f"%{term}%"), 150), else_=0)
                    + case((name_l.like(f"%{term}%"), 120), else_=0)
                    + case((source_l.like(f"%{term}%"), 70), else_=0)
                    + case((category_l.like(f"%{term}%"), 60), else_=0)
                )

                for tok in tokens:
                    tok_like = f"%{tok}%"
                    rank_expr = (
                        rank_expr
                        + case((code_l.like(tok_like), 70), else_=0)
                        + case((name_l.like(tok_like), 45), else_=0)
                        + case((source_l.like(tok_like), 25), else_=0)
                        + case((category_l.like(tok_like), 20), else_=0)
                        + case((provider_l.like(tok_like), 18), else_=0)
                        + case((asset_class_l.like(tok_like), 15), else_=0)
                        + case((country_l.like(tok_like), 12), else_=0)
                    )
                    if len(tok) >= 3:
                        rank_expr = rank_expr + case(
                            (source_code_l.like(tok_like), 10), else_=0
                        )

                timeseries_query = timeseries_query.order_by(
                    rank_expr.desc(),
                    Timeseries.favorite.desc(),
                    Timeseries.code.asc(),
                )

    # Apply filters
    if category:
        timeseries_query = timeseries_query.filter(Timeseries.category == category)
    if asset_class:
        timeseries_query = timeseries_query.filter(
            Timeseries.asset_class == asset_class
        )
    if provider:
        timeseries_query = timeseries_query.filter(Timeseries.provider == provider)

    # Apply pagination
    if offset:
        timeseries_query = timeseries_query.offset(offset)
    if limit:
        timeseries_query = timeseries_query.limit(limit)

    timeseries_list = timeseries_query.all()

    formatted_timeseries = []
    for ts in timeseries_list:
        formatted_ts = TimeseriesResponse(
            id=str(ts.id),
            code=str(ts.code) if ts.code else None,
            name=str(ts.name) if ts.name else None,
            provider=ts.provider,
            asset_class=str(ts.asset_class) if ts.asset_class else None,
            category=ts.category,
            start=ts.start,
            end=ts.end,
            num_data=int(ts.num_data) if ts.num_data is not None else None,
            source=ts.source,
            source_code=getattr(ts, "source_code", None),
            frequency=str(ts.frequency) if ts.frequency else None,
            unit=getattr(ts, "unit", None),
            scale=getattr(ts, "scale", None),
            currency=getattr(ts, "currency", None),
            country=getattr(ts, "country", None),
            remark=str(ts.remark) if ts.remark else None,
            favorite=bool(ts.favorite) if ts.favorite is not None else False,
        )
        formatted_timeseries.append(formatted_ts)

    logger.info("Retrieved %d timeseries records", len(formatted_timeseries))
    return formatted_timeseries


@router.get("/timeseries/sources")
def get_timeseries_sources(
    db: SessionType = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """Return distinct source names for timeseries that have source_code set."""
    from sqlalchemy import distinct

    ensure_connection()
    rows = (
        db.query(distinct(Timeseries.source))
        .filter(Timeseries.source_code.isnot(None))
        .filter(Timeseries.source.isnot(None))
        .all()
    )
    return sorted([r[0] for r in rows])


# Universal Excel formula template for the download endpoint.
# Matches user's EXACT formatted string (including spaces).
# __C__ is replaced with the actual column letter (B, C, D, ...) per timeseries.
_DOWNLOAD_FORMULA_TEMPLATE = (
    """=IF(__C__6="FactSet",IF(ISNUMBER(SEARCH(__C__3, "FDS_ECON_DATA; FDS_COM_DATA",1)),"""
    """FDSC("","","PSETCAL(SEVENDAY);"&__C__3&"('" & __C__2 & "'," & $C$1+2 & "," & $B$1 & ", D, NONE, NONE)"),\n"""
    """FDSC("","","PSETCAL(SEVENDAY);NO_REPEAT_F(SPEC_ID_DATA('" & __C__2 & ":" & __C__3 & "','" & $C$1+2 & "','" & $B$1 & "', D, NONE, NONE,2))")),\n"""
    """IF(__C__6="Bloomberg",BDH(__C__2,__C__3,$B$1,$C$1+2,"SORT", "TRUE","DTS", "FALSE","DAYS", "C","FILL", "B"),\n"""
    """IF(__C__6="Infomax",IMDH(__C__5,__C__2&"",__C__3,$B$1,$C$1,9999,"Per=일,sort=D,real=false,Bizday=12,Quote=종가,ROUND=9,Pos=20,Orient=V,Title="&__C__7&",DtFmt=1,TmFmt=1,unit=true"),\n"""
    """NA())))"""
)


@router.get("/timeseries/download_template")
def download_template(
    source: Optional[List[str]] = Query(
        None,
        description="Filter by source(s): Bloomberg, FactSet, Infomax (repeatable)",
    ),
    start_date: Optional[str] = Query(
        None, description="Start date (ISO-8601), default ~2 years ago"
    ),
    end_date: Optional[str] = Query(
        None, description="End date (ISO-8601), default today"
    ),
    db: SessionType = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Generate an Excel template for data download/upload workflow.

    Produces a .xlsx file with metadata rows and pre-built formulas
    matching the Google Sheets Bloomberg/FactSet/Infomax workflow.
    Accepts multiple source values via repeated query params.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font

    ensure_connection()

    # Date range
    end_dt = pd.to_datetime(end_date) if end_date else pd.Timestamp.now()
    start_dt = (
        pd.to_datetime(start_date) if start_date else end_dt - pd.DateOffset(years=1)
    )
    start_str = start_dt.strftime("%Y-%m-%d")
    end_str = end_dt.strftime("%Y-%m-%d")

    # Query timeseries that have source_code set
    query = db.query(Timeseries).filter(Timeseries.source_code.isnot(None))
    if source:
        query = query.filter(Timeseries.source.in_(source))
    all_ts = query.order_by(Timeseries.source, Timeseries.code).all()

    if not all_ts:
        raise HTTPException(
            status_code=404, detail="No timeseries with source_code found."
        )

    # Group by source
    grouped: Dict[str, list] = {}
    for ts in all_ts:
        src = ts.source or "Unknown"
        grouped.setdefault(src, []).append(ts)

    wb = Workbook()
    wb.remove(wb.active)

    # Styles — white background, black text
    header_font = Font(name="Consolas", size=9, color="000000")
    value_font = Font(name="Consolas", size=9, color="000000")
    code_font = Font(name="Consolas", size=9, color="000000", bold=True)
    formula_font = Font(name="Consolas", size=8, color="000000", italic=True)
    date_font = Font(name="Consolas", size=9, color="000000")

    # Date range extent (for formula row count)
    date_top = end_dt + pd.DateOffset(days=2)

    for sheet_source, ts_list in grouped.items():
        ws = wb.create_sheet(title=(sheet_source or "Unknown")[:31])

        # Row 1: Header info — matches Sample.xlsx layout (A1=source, B1=start, C1=end)
        ws.cell(row=1, column=1, value=sheet_source).font = Font(
            name="Consolas", size=10, color="0000FF", bold=True
        )
        # B1 = start date, C1 = end date
        ws.cell(row=1, column=2, value=start_dt.to_pydatetime()).font = value_font
        ws.cell(row=1, column=2).number_format = "mm-dd-yy"
        ws.cell(row=1, column=3, value=end_dt.to_pydatetime()).font = value_font
        ws.cell(row=1, column=3).number_format = "mm-dd-yy"

        # Metadata row labels (col A)
        row_labels = {
            2: "source_ticker",
            3: "source_field",
            4: "source_code",
            5: "asset_class",
            6: "source",
            7: "name",
            8: "code",
        }
        for r, label in row_labels.items():
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = header_font

        # Fill columns B+ with timeseries metadata
        for col_idx, ts in enumerate(ts_list, start=2):
            sc = str(ts.source_code) if ts.source_code else ":"
            parts = sc.rsplit(":", 1) if ":" in sc else [sc, ""]
            ticker = parts[0] if len(parts) > 0 else ""
            field = parts[1] if len(parts) > 1 else ""

            c = ws.cell(row=2, column=col_idx, value=ticker)
            c.font = value_font
            c.number_format = '@'
            c.data_type = 's'  # Force string — preserves leading zeros like "001"
            c = ws.cell(row=3, column=col_idx, value=field)
            c.font = value_font
            c.number_format = '@'
            c.data_type = 's'
            c = ws.cell(row=4, column=col_idx, value=sc)
            c.font = value_font
            c.number_format = '@'
            c.data_type = 's'
            ws.cell(row=5, column=col_idx, value=ts.asset_class or "").font = value_font
            ws.cell(row=6, column=col_idx, value=ts.source or "").font = value_font
            ws.cell(row=7, column=col_idx, value=ts.name or "").font = value_font
            c = ws.cell(row=8, column=col_idx, value=ts.code or "")
            c.font = code_font
            c.number_format = '@'
            c.data_type = 's'

            # Universal formula — column letter swapped via template
            from openpyxl.utils import get_column_letter

            col = get_column_letter(col_idx)
            formula_text = _DOWNLOAD_FORMULA_TEMPLATE.replace("__C__", col)

            ws.cell(row=9, column=col_idx, value=formula_text).font = formula_font

        # Date column (A9+) — descending from end_date+2 toward start_date
        num_date_rows = (date_top - start_dt).days + 1
        # A9 = C1+2 (most recent), then A10 = A9-1, A11 = A10-1, etc.
        c = ws.cell(row=9, column=1, value="=C1+2")
        c.font = date_font
        c.number_format = "mm-dd-yy"
        for i in range(1, num_date_rows):
            c = ws.cell(row=9 + i, column=1, value=f"=A{9 + i - 1}-1")
            c.font = date_font
            c.number_format = "mm-dd-yy"

        # Column widths
        ws.column_dimensions["A"].width = 16
        for col_idx in range(2, len(ts_list) + 2):
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = (
                14
            )

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"timeseries_template_{end_dt.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/timeseries/upload_template_data")
@_limiter.limit("10/minute")
async def upload_template_data(
    request: Request,
    file: UploadFile = File(...),
    _current_user: User = Depends(get_current_admin_user),
):
    """Upload filled Excel template. Same behaviour as upload_data_columnar:
    Server → saves to R2.  Local → merges into local DB + cloud DB.
    """
    import asyncio

    # Validate file type
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
    }
    if file.content_type and file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="File must be an Excel spreadsheet (.xlsx)")

    contents = await file.read()
    if len(contents) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File exceeds 50MB size limit")

    # Run blocking work (openpyxl + DB) in threadpool — fresh session inside thread
    result = await asyncio.to_thread(_process_template_upload, contents)
    return result


def _process_template_upload(contents: bytes):
    """Blocking worker: parse Excel template and merge into DB (same as upload_data_columnar)."""
    import io
    import math
    import time as _time
    from datetime import datetime as _dt, date as _date
    from openpyxl import load_workbook
    from ix.db.boto import Boto
    from ix.misc.settings import Settings

    _t0 = _time.time()
    logger.info("Template upload: parsing %d bytes...", len(contents))
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel format: {e}")
    logger.info("Template upload: workbook loaded in %.1fs", _time.time() - _t0)

    all_records: dict = {}  # {code: {date_str: value}}
    for ws in wb.worksheets:
        row_num = 0
        codes_by_col_idx: dict = {}  # {0-based col index: code}
        for row in ws.iter_rows(values_only=True):
            row_num += 1
            if row_num < 8:
                continue
            if row_num == 8:
                # Row 8 = codes (skip col A at index 0)
                for ci, val in enumerate(row):
                    if ci > 0 and val:
                        codes_by_col_idx[ci] = str(val).strip()
                continue
            # Rows 9+ = data
            if not row or not codes_by_col_idx:
                continue
            date_cell = row[0]
            if date_cell is None:
                continue
            date_str = None
            if isinstance(date_cell, (_dt, _date)):
                date_str = date_cell.strftime("%Y-%m-%d")
            elif isinstance(date_cell, (int, float)):
                try:
                    date_str = pd.to_datetime(date_cell, unit="D", origin="1899-12-30").strftime("%Y-%m-%d")
                except Exception:
                    pass
            elif isinstance(date_cell, str):
                try:
                    date_str = pd.to_datetime(date_cell).strftime("%Y-%m-%d")
                except Exception:
                    pass
            if not date_str:
                continue
            for ci, code in codes_by_col_idx.items():
                if ci < len(row) and row[ci] is not None:
                    try:
                        val = float(row[ci])
                        if math.isnan(val):
                            continue
                    except Exception:
                        continue
                    if code not in all_records:
                        all_records[code] = {}
                    all_records[code][date_str] = val
    wb.close()
    logger.info("Template upload: parsed %d codes in %.1fs", len(all_records), _time.time() - _t0)

    if not all_records:
        raise HTTPException(status_code=400, detail="No data found in uploaded file.")

    # Build columnar format
    all_dates = sorted({d for rec in all_records.values() for d in rec})
    columns = {}
    for code, rec in all_records.items():
        columns[code] = [rec.get(d) for d in all_dates]

    num_dates = len(all_dates)
    num_cols = len(columns)
    codes = sorted(columns.keys())

    # Server: save to R2
    if Settings.is_server:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"uploads/{timestamp}_col_{num_cols}x{num_dates}.json"
        try:
            Boto().save_json(
                {"format": "columnar", "dates": all_dates, "columns": columns},
                filename,
            )
        except Exception as e:
            logger.exception("Failed to upload to R2: %s", e)
            raise HTTPException(status_code=500, detail="Failed to save to storage")
        return {
            "message": f"Saved {num_dates} dates × {num_cols} columns to storage.",
            "file": filename,
            "codes": codes,
        }

    # Local: merge into local DB + cloud DB (fresh session for thread safety)
    ensure_connection()
    dates_index = pd.to_datetime(all_dates, errors="coerce")
    df = pd.DataFrame(columns, index=dates_index)
    df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)
    if df.empty:
        raise HTTPException(status_code=400, detail="No valid records after cleaning.")

    logger.info("Template upload: merging %d codes into local DB...", len(df.columns))
    _t1 = _time.time()
    with Session() as local_db:
        result = _merge_columnar_to_db(df, local_db)
    logger.info("Template upload: local DB merge done in %.1fs (%d updated)", _time.time() - _t1, len(result["updated"]))
    response = {
        "message": f"Merged {result['points']} points for {len(result['updated'])} codes.",
        "db_updated": result["updated"],
        "db_points_merged": result["points"],
    }
    if result["not_found"]:
        response["warning"] = f"Codes not found in database: {result['not_found']}"

    # Cloud sync in background thread — don't block the response
    import threading
    def _bg_cloud_sync(df_copy, resp_copy):
        try:
            _sync_to_cloud(df_copy, resp_copy)
            logger.info("Template upload: cloud sync done in background (%.1fs total)", _time.time() - _t0)
        except Exception as e:
            logger.exception("Template upload: cloud sync failed: %s", e)
    threading.Thread(target=_bg_cloud_sync, args=(df.copy(), {}), daemon=True).start()
    response["cloud_status"] = "syncing in background"

    logger.info("Template upload: returning response after %.1fs", _time.time() - _t0)
    return response


@router.post("/timeseries")
@_limiter.limit("10/minute")
def create_or_update_timeseries_bulk(
    request: Request,
    payload: List[TimeseriesCreate],
    db: SessionType = Depends(get_db),
    _current_user: User = Depends(get_current_admin_user),
):
    """
    POST /api/timeseries - Create new or update existing timeseries metadata (bulk).

    Request body: List of timeseries objects. Each object can have:
    - id: Optional UUID string to update existing timeseries by ID
    - code: Required for new timeseries, optional if id is provided
    - Other fields: name, provider, asset_class, category, etc.

    Update logic:
    1. If 'id' is provided, find by ID first
    2. If not found by ID and 'code' is provided, find by code
    3. If not found, create new (requires 'code')
    """
    ensure_connection()

    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload provided.")

    updated_codes = []
    created_codes = []
    errors = []

    # Pre-fetch existing timeseries
    payload_ids = [str(ts.id) for ts in payload if ts.id]
    payload_codes = [ts.code for ts in payload if ts.code]

    existing_by_id = {}
    if payload_ids:
        for ts in db.query(Timeseries).filter(Timeseries.id.in_(payload_ids)).all():
            existing_by_id[str(ts.id)] = ts

    existing_by_code = {}
    if payload_codes:
        for ts in db.query(Timeseries).filter(Timeseries.code.in_(payload_codes)).all():
            existing_by_code[str(ts.code)] = ts

    for ts_data in payload:
        try:
            ts = None

            # If ID is provided, try to find by ID first
            if ts_data.id:
                ts = existing_by_id.get(str(ts_data.id))
                if ts:
                    updated_codes.append(ts.code if ts.code else str(ts_data.id))

            # If not found by ID, try to find by code
            if ts is None and ts_data.code:
                ts = existing_by_code.get(str(ts_data.code))
                if ts:
                    updated_codes.append(ts_data.code)

            # If still not found, create new (requires code)
            if ts is None:
                if not ts_data.code:
                    errors.append(
                        "Either 'id' or 'code' is required to create/update timeseries"
                    )
                    continue
                # Create new timeseries
                ts = Timeseries(code=ts_data.code)
                db.add(ts)
                existing_by_code[str(ts_data.code)] = (
                    ts  # Register for subsequent lookups
                )
                created_codes.append(ts_data.code)

            # Update fields
            # Update code if provided (only if different from current)
            if ts_data.code and ts_data.code != ts.code:
                ts.code = ts_data.code

            if ts_data.name is not None:
                ts.name = str(ts_data.name)[:200] if ts_data.name else None
            if ts_data.provider is not None:
                ts.provider = str(ts_data.provider)[:100] if ts_data.provider else None
            if ts_data.asset_class is not None:
                ts.asset_class = (
                    str(ts_data.asset_class)[:50] if ts_data.asset_class else None
                )
            if ts_data.category is not None:
                ts.category = str(ts_data.category)[:100] if ts_data.category else None
            if ts_data.source is not None:
                ts.source = str(ts_data.source)[:100] if ts_data.source else None
            if ts_data.source_code is not None:
                ts.source_code = (
                    str(ts_data.source_code)[:2000] if ts_data.source_code else None
                )
            if ts_data.frequency is not None:
                ts.frequency = (
                    str(ts_data.frequency)[:20] if ts_data.frequency else None
                )
            if ts_data.unit is not None:
                ts.unit = str(ts_data.unit)[:50] if ts_data.unit else None
            if ts_data.scale is not None:
                ts.scale = ts_data.scale
            if ts_data.currency is not None:
                ts.currency = str(ts_data.currency)[:10] if ts_data.currency else None
            if ts_data.country is not None:
                ts.country = str(ts_data.country)[:100] if ts_data.country else None
            if ts_data.remark is not None:
                ts.remark = str(ts_data.remark) if ts_data.remark else None
            if ts_data.favorite is not None:
                ts.favorite = ts_data.favorite

            # We use a nested transaction (SAVEPOINT) to safely catch individual loop errors
            # without breaking the entire batch transaction.
            try:
                with db.begin_nested():
                    db.flush()
            except Exception as nested_e:
                raise nested_e

        except Exception as e:
            identifier = ts_data.id or ts_data.code or "unknown"
            logger.error("Error updating timeseries %s: %s", identifier, e)
            errors.append(f"Error updating {identifier}: {str(e)}")
            continue

    try:
        db.commit()
    except Exception as e:
        logger.error("Error committing bulk update: %s", e)
        db.rollback()
        errors.append(f"Transaction failed: {str(e)}")

    response = {
        "message": f"Successfully processed {len(payload)} timeseries objects.",
        "created_codes": created_codes,
        "created_count": len(created_codes),
        "updated_codes": updated_codes,
        "updated_count": len(updated_codes),
    }

    if errors:
        response["errors"] = errors
        response["error_count"] = len(errors)

    return response


@router.put("/timeseries/{code}", response_model=TimeseriesResponse)
@_limiter.limit("30/minute")
def update_timeseries(
    request: Request,
    code: str,
    payload: TimeseriesUpdate,
    db: SessionType = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """
    PUT /api/timeseries/{code} - Update a single timeseries metadata entry.

    This endpoint follows REST best practices by using PUT for updates.
    Only provided fields are updated (no payload -> 400).
    """
    ensure_connection()

    ts = db.query(Timeseries).filter(Timeseries.code == code).first()
    if not ts:
        raise HTTPException(
            status_code=404, detail=f"Timeseries with code '{code}' not found"
        )

    update_fields = payload.model_dump(exclude_unset=True)
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields provided to update.")

    try:
        # Apply changes
        try:
            _apply_timeseries_updates(ts, update_fields)
        except (TypeError, ValueError):
            raise HTTPException(
                status_code=400, detail="Invalid field type in payload."
            )

        ts.updated = datetime.now()
        db.commit()
        db.refresh(ts)

        return TimeseriesResponse(
            id=str(ts.id),
            code=ts.code,
            name=ts.name,
            provider=ts.provider,
            asset_class=ts.asset_class,
            category=ts.category,
            start=ts.start,
            end=ts.end,
            num_data=ts.num_data,
            source=ts.source,
            source_code=getattr(ts, "source_code", None),
            frequency=ts.frequency,
            unit=getattr(ts, "unit", None),
            scale=getattr(ts, "scale", None),
            currency=getattr(ts, "currency", None),
            country=getattr(ts, "country", None),
            remark=getattr(ts, "remark", None),
            favorite=getattr(ts, "favorite", False),
        )
    except HTTPException:
        # Already constructed, just re-raise
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("Failed to update timeseries %s: %s", code, e)
        raise HTTPException(
            status_code=500, detail=f"Failed to update timeseries '{code}'"
        )


@router.post("/timeseries/{code}", response_model=TimeseriesResponse, status_code=201)
@_limiter.limit("10/minute")
def create_timeseries(
    request: Request,
    code: str,
    payload: TimeseriesCreate,
    db: SessionType = Depends(get_db),
    _current_user: User = Depends(get_current_admin_user),
):
    """
    POST /api/timeseries/{code} - Create a single timeseries metadata entry.

    - Returns 409 if the code already exists.
    - Body fields mirror the bulk create schema; path code overrides body code.
    """
    ensure_connection()

    # Validate existence
    existing = db.query(Timeseries).filter(Timeseries.code == code).first()
    if existing:
        raise HTTPException(
            status_code=409, detail=f"Timeseries with code '{code}' already exists"
        )

    # Build data from payload with path code taking precedence
    create_data = payload.model_dump(exclude_unset=True)
    create_data["code"] = code  # enforce path code

    try:
        ts = Timeseries(code=code)
        _apply_timeseries_updates(ts, create_data)
        ts.created = datetime.now()
        ts.updated = datetime.now()

        db.add(ts)
        db.commit()
        db.refresh(ts)

        return TimeseriesResponse(
            id=str(ts.id),
            code=ts.code,
            name=ts.name,
            provider=ts.provider,
            asset_class=ts.asset_class,
            category=ts.category,
            start=ts.start,
            end=ts.end,
            num_data=ts.num_data,
            source=ts.source,
            source_code=getattr(ts, "source_code", None),
            frequency=ts.frequency,
            unit=getattr(ts, "unit", None),
            scale=getattr(ts, "scale", None),
            currency=getattr(ts, "currency", None),
            country=getattr(ts, "country", None),
            remark=getattr(ts, "remark", None),
            favorite=getattr(ts, "favorite", False),
        )
    except Exception as e:
        db.rollback()
        logger.error("Failed to create timeseries %s: %s", code, e)
        raise HTTPException(
            status_code=500, detail=f"Failed to create timeseries '{code}'"
        )


@router.delete("/timeseries/{code}", status_code=status.HTTP_204_NO_CONTENT)
@_limiter.limit("10/minute")
def delete_timeseries(
    request: Request,
    code: str,
    db: SessionType = Depends(get_db),
    current_user: User = Depends(get_current_admin_user),
):
    """
    DELETE /api/timeseries/{code} - Delete a timeseries by code.

    Admin-only endpoint.
    """
    ensure_connection()

    ts = db.query(Timeseries).filter(Timeseries.code == code).first()
    if not ts:
        raise HTTPException(
            status_code=404, detail=f"Timeseries with code '{code}' not found"
        )

    try:
        db.delete(ts)
        db.commit()
        logger.info("Admin %s deleted timeseries: %s", current_user.email, code)
        return
    except Exception as e:
        db.rollback()
        logger.error("Failed to delete timeseries %s: %s", code, e)
        raise HTTPException(
            status_code=500, detail=f"Failed to delete timeseries '{code}'"
        )


@router.get("/timeseries/{timeseries_id}", response_model=TimeseriesResponse)
def get_timeseries_by_id(
    timeseries_id: str,
    db: SessionType = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """
    GET /api/timeseries/{id} - Get detailed timeseries information by ID (code) with full data.
    """
    ensure_connection()

    ts = db.query(Timeseries).filter(Timeseries.code == timeseries_id).first()

    if not ts:
        raise HTTPException(status_code=404, detail="Timeseries not found")

    return TimeseriesResponse(
        id=str(ts.id),
        code=ts.code,
        name=ts.name,
        provider=ts.provider,
        asset_class=ts.asset_class,
        category=ts.category,
        start=ts.start,
        end=ts.end,
        num_data=ts.num_data,
        source=ts.source,
        source_code=getattr(ts, "source_code", None),
        frequency=ts.frequency,
        unit=getattr(ts, "unit", None),
        scale=getattr(ts, "scale", None),
        currency=getattr(ts, "currency", None),
        country=getattr(ts, "country", None),
        remark=getattr(ts, "remark", None),
        favorite=getattr(ts, "favorite", False),
    )


@router.get("/timeseries/code/{code}", response_model=TimeseriesResponse)
def get_timeseries_by_code(
    code: str,
    db: SessionType = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """
    GET /api/timeseries/code/{code} - Get timeseries by its code.
    """
    ensure_connection()

    ts = db.query(Timeseries).filter(Timeseries.code == code).first()

    if not ts:
        raise HTTPException(
            status_code=404, detail=f"Timeseries with code '{code}' not found"
        )

    return TimeseriesResponse(
        id=str(ts.id),
        code=ts.code,
        name=ts.name,
        provider=ts.provider,
        asset_class=ts.asset_class,
        category=ts.category,
        start=ts.start,
        end=ts.end,
        num_data=ts.num_data,
        source=ts.source,
        source_code=getattr(ts, "source_code", None),
        frequency=ts.frequency,
        unit=getattr(ts, "unit", None),
        scale=getattr(ts, "scale", None),
        currency=getattr(ts, "currency", None),
        country=getattr(ts, "country", None),
        remark=getattr(ts, "remark", None),
        favorite=getattr(ts, "favorite", False),
    )


@router.get("/timeseries/favorites")
def get_favorite_timeseries_data(
    start_date: Optional[str] = Query(None, description="Start date (ISO-8601)"),
    end_date: Optional[str] = Query(None, description="End date (ISO-8601)"),
    db: SessionType = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """
    GET /api/timeseries/favorites - Get all favorite timeseries data as a concatenated DataFrame.

    Returns column-oriented JSON format with dates sorted descending.
    """
    ensure_connection()

    try:
        # Query favorite timeseries metadata only — data_record loaded lazily
        # per item via _get_or_create_data_record() to avoid a single massive
        # JOIN that pulls all JSONB payloads into memory at once.
        favorite_timeseries = (
            db.query(Timeseries)
            .filter(Timeseries.favorite == True)
            .all()
        )

        if not favorite_timeseries:
            return Response(
                content=json.dumps({"Date": []}, ensure_ascii=False),
                media_type="application/json",
            )

        # Collect all series data as pandas Series for concatenation
        series_list = []

        for ts in favorite_timeseries:
            try:
                ts_code = ts.code
                data_record = ts._get_or_create_data_record(db)
                column_data = (
                    data_record.data if data_record and data_record.data else {}
                )
                frequency = ts.frequency

                # Convert JSONB dict to pandas Series
                if column_data and isinstance(column_data, dict):
                    ts_data = pd.Series(column_data)
                    try:
                        ts_data.index = pd.to_datetime(ts_data.index)
                        ts_data = pd.to_numeric(ts_data, errors="coerce").dropna()
                        ts_data.name = ts_code
                        if frequency and len(ts_data) > 0:
                            ts_data = (
                                ts_data.sort_index()
                                .resample(str(frequency))
                                .last()
                                .dropna()
                            )
                        else:
                            ts_data = ts_data.sort_index()

                        if not ts_data.empty:
                            series_list.append(ts_data)
                    except Exception:
                        try:
                            valid_dates = pd.to_datetime(ts_data.index, errors="coerce")
                            ts_data = ts_data[valid_dates.notna()]
                            ts_data.index = pd.to_datetime(ts_data.index)
                            ts_data = pd.to_numeric(ts_data, errors="coerce").dropna()
                            ts_data.name = ts_code
                            ts_data = ts_data.sort_index()
                            if not ts_data.empty:
                                series_list.append(ts_data)
                        except Exception:
                            logger.warning(
                                f"Error processing timeseries {ts_code}: could not convert to Series"
                            )
                            continue
            except Exception as e:
                logger.warning("Error processing favorite timeseries %s: %s", ts.code, e)
                continue

        # Concatenate all series into a DataFrame
        if series_list:
            df = pd.concat(series_list, axis=1)
            df.index.name = "Date"

            # Optional date slicing
            start_ts = (
                pd.to_datetime(start_date, errors="coerce") if start_date else None
            )
            end_ts = pd.to_datetime(end_date, errors="coerce") if end_date else None

            # Normalize timezone info
            try:
                df.index = pd.DatetimeIndex(df.index).tz_localize(None)
            except Exception:
                try:
                    df.index = pd.DatetimeIndex(df.index).tz_convert(None)
                except Exception:
                    pass

            df = df.resample("D").last()

            if isinstance(start_ts, pd.Timestamp):
                try:
                    start_ts = start_ts.tz_localize(None)
                except Exception:
                    try:
                        start_ts = start_ts.tz_convert(None)
                    except Exception:
                        pass

            if isinstance(end_ts, pd.Timestamp):
                try:
                    end_ts = end_ts.tz_localize(None)
                except Exception:
                    try:
                        end_ts = end_ts.tz_convert(None)
                    except Exception:
                        pass

            # Apply slicing if bounds are valid
            if isinstance(start_ts, pd.Timestamp):
                df = df[df.index >= start_ts]
            if isinstance(end_ts, pd.Timestamp):
                df = df[df.index <= end_ts]

            # Sort dates descending
            df = df.sort_index(ascending=True)

            # Convert to column-oriented format
            df_indexed = df.reset_index()
            column_dict = OrderedDict()

            for col in df_indexed.columns:
                values = df_indexed[col].tolist()
                cleaned_values = []
                for v in values:
                    if v is None or (isinstance(v, float) and math.isnan(v)):
                        cleaned_values.append(None)
                    elif isinstance(v, (pd.Timestamp, datetime)):
                        cleaned_values.append(v.isoformat())
                    else:
                        cleaned_values.append(v)
                column_dict[col] = cleaned_values

            return Response(
                content=json.dumps(column_dict, ensure_ascii=False),
                media_type="application/json",
            )
        else:
            return Response(
                content=json.dumps({"Date": []}, ensure_ascii=False),
                media_type="application/json",
            )

    except Exception as e:
        logger.exception("Error retrieving favorite timeseries: %s", e)
        raise HTTPException(status_code=500, detail="Internal server error")


async def _parse_codes_from_request(request: Request) -> List[str]:
    """Parse codes from request body, query parameters, or X-Codes header."""
    codes = []

    # Try to get from request body (works for POST, and GET with body in some clients)
    try:
        body = await request.json()
        if isinstance(body, list):
            codes = [str(c).strip() for c in body if str(c).strip()]
        elif isinstance(body, dict) and body:
            raw_codes = body.get("codes")
            if isinstance(raw_codes, list):
                codes = [str(c).strip() for c in raw_codes if str(c).strip()]
            elif isinstance(raw_codes, str):
                codes = [c.strip() for c in raw_codes.split(",") if c.strip()]
    except Exception:
        pass

    # Fallback to query parameters (for GET requests)
    if not codes:
        query_codes = request.query_params.get("codes")
        if query_codes:
            codes = [c.strip() for c in query_codes.split(",") if c.strip()]

    # Fallback to header if body and query didn't work
    if not codes:
        header_codes = request.headers.get("X-Codes")
        if header_codes:
            # Try to parse as JSON first (handles codes with commas and special chars)
            try:
                parsed = json.loads(header_codes)
                if isinstance(parsed, list):
                    codes = [str(c).strip() for c in parsed if str(c).strip()]
                    logger.debug(
                        f"Parsed {len(codes)} codes from JSON array in X-Codes header"
                    )
                elif isinstance(parsed, dict) and "codes" in parsed:
                    codes_list = parsed["codes"]
                    if isinstance(codes_list, list):
                        codes = [str(c).strip() for c in codes_list if str(c).strip()]
                        logger.debug(
                            f"Parsed {len(codes)} codes from JSON dict in X-Codes header"
                        )
                    elif isinstance(codes_list, str):
                        codes = [c.strip() for c in codes_list.split(",") if c.strip()]
                        logger.warning(
                            "X-Codes header contains dict with string 'codes' - using comma-split (may break)"
                        )
                else:
                    logger.warning(
                        f"X-Codes JSON parsed but unexpected format: {type(parsed)}"
                    )
            except (json.JSONDecodeError, ValueError) as e:
                # Check if header looks like it might contain complex expressions with commas
                # If it contains function calls, quotes, or parentheses, it's likely JSON should be used
                has_complex_expressions = any(
                    char in header_codes
                    for char in [
                        "(",
                        ")",
                        "{",
                        "}",
                        '"',
                        "'",
                        "Series(",
                        "MultiSeries(",
                    ]
                )

                if has_complex_expressions:
                    logger.error(
                        f"X-Codes header contains complex expressions but is not valid JSON. "
                        f'Please send codes as JSON array: ["code1", "code2"]. '
                        f"Error: {e}. Header preview: {header_codes[:200]}..."
                    )
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "X-Codes header must be a JSON array when codes contain commas or special characters. "
                            'Example: X-Codes: ["CODE1", "Series(\'CODE2\').pct_change()"]'
                        ),
                    )

                # Fallback to comma-separated for simple codes only
                logger.warning(
                    f"X-Codes header is not valid JSON (falling back to comma-separated): {e}. "
                    f"Header value starts with: {header_codes[:100]}..."
                )
                codes = [c.strip() for c in header_codes.split(",") if c.strip()]

    if not codes:
        raise HTTPException(
            status_code=400,
            detail="No codes provided. Provide JSON body with 'codes' list, 'codes' query parameter, or 'X-Codes' header.",
        )

    # Filter out reserved column names
    RESERVED_NAMES = {"Date", "date"}  # Reserved for index column
    codes = [c for c in codes if c not in RESERVED_NAMES]

    if not codes:
        raise HTTPException(
            status_code=400,
            detail="No valid codes provided after filtering reserved names.",
        )

    # Deduplicate while preserving order
    seen_codes = set()
    unique_codes = []
    for code in codes:
        if code not in seen_codes:
            unique_codes.append(code)
            seen_codes.add(code)

    return unique_codes


def _process_database_timeseries(
    ts: Timeseries, db: SessionType
) -> Optional[pd.Series]:
    """Process a timeseries from the database and return as pandas Series."""
    try:
        data_record = ts._get_or_create_data_record(db)
        column_data = data_record.data if data_record and data_record.data else {}
        frequency = ts.frequency

        if column_data and isinstance(column_data, dict):
            ts_data = pd.Series(column_data)
            try:
                ts_data.index = pd.to_datetime(ts_data.index)
                ts_data = pd.to_numeric(ts_data, errors="coerce").dropna()
                ts_data.name = ts.code
                if frequency and len(ts_data) > 0:
                    ts_data = (
                        ts_data.sort_index().resample(str(frequency)).last().dropna()
                    )
                else:
                    ts_data = ts_data.sort_index()

                if not ts_data.empty:
                    return ts_data
            except Exception:
                try:
                    valid_dates = pd.to_datetime(ts_data.index, errors="coerce")
                    ts_data = ts_data[valid_dates.notna()]
                    ts_data.index = pd.to_datetime(ts_data.index)
                    ts_data = pd.to_numeric(ts_data, errors="coerce").dropna()
                    ts_data.name = ts.code
                    ts_data = ts_data.sort_index()
                    if not ts_data.empty:
                        return ts_data
                except Exception:
                    logger.warning(
                        f"Error processing timeseries {ts.code}: could not convert to Series"
                    )
                    return None
    except Exception as e:
        logger.warning("Error processing database timeseries %s: %s", ts.code, e)
        return None

    return None


def _evaluate_expression(
    code: str, start_date: Optional[str], end_date: Optional[str]
) -> List[pd.Series]:
    """Evaluate code as Python expression and return list of Series."""
    try:
        logger.info(
            "Code %s not found in database, attempting to evaluate as expression", code
        )
        evaluated_series = safe_eval_expression(code, TIMESERIES_EXPRESSION_CONTEXT)
        series_list = []

        # Handle both Series and DataFrame results
        if isinstance(evaluated_series, pd.Series):
            evaluated_series.name = code
            if not evaluated_series.empty:
                # Apply date filtering if provided
                if start_date:
                    start_dt = pd.to_datetime(start_date, errors="coerce")
                    if start_dt:
                        evaluated_series = evaluated_series.loc[
                            evaluated_series.index >= start_dt
                        ]
                if end_date:
                    end_dt = pd.to_datetime(end_date, errors="coerce")
                    if end_dt:
                        evaluated_series = evaluated_series.loc[
                            evaluated_series.index <= end_dt
                        ]
                series_list.append(evaluated_series)
        elif isinstance(evaluated_series, pd.DataFrame):
            # If DataFrame, convert each column to a series
            for col in evaluated_series.columns:
                col_series = evaluated_series[col].copy()
                col_series.name = col
                # Apply date filtering if provided
                if start_date:
                    start_dt = pd.to_datetime(start_date, errors="coerce")
                    if start_dt:
                        col_series = col_series.loc[col_series.index >= start_dt]
                if end_date:
                    end_dt = pd.to_datetime(end_date, errors="coerce")
                    if end_dt:
                        col_series = col_series.loc[col_series.index <= end_dt]
                if not col_series.empty:
                    series_list.append(col_series)
        else:
            logger.warning(
                "Evaluated expression %s did not return a Series or DataFrame", code
            )

        return series_list
    except UnsafeExpressionError as e:
        logger.warning("Rejected custom timeseries expression %s: %s", code, e)
        return []
    except Exception as e:
        logger.warning(
            "Code %s not found in database and failed to evaluate as expression: %s", code, e
        )
        return []


def _normalize_timezone(ts: pd.Timestamp) -> pd.Timestamp:
    """Normalize timezone info from a Timestamp."""
    try:
        return ts.tz_localize(None)
    except Exception:
        try:
            return ts.tz_convert(None)
        except Exception:
            return ts


def _format_dataframe_response(
    df: pd.DataFrame,
    series_list: List[pd.Series],
    start_date: Optional[str],
    end_date: Optional[str],
) -> Response:
    """Format DataFrame into column-oriented JSON response."""
    df.index.name = "Date"

    # Optional date bounds
    start_ts = pd.to_datetime(start_date, errors="coerce") if start_date else None
    end_ts = pd.to_datetime(end_date, errors="coerce") if end_date else None

    # Normalize timezone info
    try:
        df.index = pd.DatetimeIndex(df.index).tz_localize(None)
    except Exception:
        try:
            df.index = pd.DatetimeIndex(df.index).tz_convert(None)
        except Exception:
            pass

    # Resample to daily frequency, drop all-NaN rows (weekends/holidays)
    df = df.resample("D").last().dropna(how="all")

    if isinstance(start_ts, pd.Timestamp):
        start_ts = _normalize_timezone(start_ts)
    if isinstance(end_ts, pd.Timestamp):
        end_ts = _normalize_timezone(end_ts)

    # Apply slicing if bounds are valid
    if isinstance(start_ts, pd.Timestamp):
        df = df[df.index >= start_ts]
    if isinstance(end_ts, pd.Timestamp):
        df = df[df.index <= end_ts]

    # Reorder columns to preserve input order
    present_in_order = [s.name for s in series_list if s.name in df.columns]
    seen_present = set()
    present_in_order = [
        c for c in present_in_order if not (c in seen_present or seen_present.add(c))
    ]
    df = df[present_in_order]

    # Sort dates ascending as requested
    df = df.sort_index(ascending=True)

    # Convert to column-oriented format
    df_indexed = df.reset_index()
    column_dict = OrderedDict()

    for col in df_indexed.columns:
        values = df_indexed[col].tolist()
        cleaned_values = []
        for v in values:
            if v is None or (isinstance(v, float) and math.isnan(v)):
                cleaned_values.append(None)
            elif isinstance(v, (pd.Timestamp, datetime)):
                cleaned_values.append(v.isoformat())
            else:
                cleaned_values.append(v)
        column_dict[col] = cleaned_values

    return Response(
        content=json.dumps(column_dict, ensure_ascii=False),
        media_type="application/json",
    )


@router.get("/timeseries.custom")
@router.post("/timeseries.custom")
@_limiter.limit("60/minute")
async def get_custom_timeseries_data(
    request: Request,
    start_date: Optional[str] = Query(None, description="Start date (ISO-8601)"),
    end_date: Optional[str] = Query(None, description="End date (ISO-8601)"),
    db: SessionType = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """
    GET/POST /api/timeseries.custom - Get selected timeseries data based on provided codes.

    Accepts JSON body (not in URL):
    { "codes": ["CODE1", "CODE2", ...] }
    or
    { "codes": "CODE1, CODE2, ..." }

    Also accepts X-Codes header:
    - JSON format (recommended for codes with commas): X-Codes: ["CODE1", "CODE2, WITH COMMA", "CODE3"]
    - Comma-separated (legacy, breaks if codes contain commas): X-Codes: CODE1, CODE2, CODE3

    Behavior:
    - Queries each code one by one from the Timeseries table
    - If code is found in database, uses stored data
    - If code is NOT found, attempts to evaluate it as a Python expression (e.g., Series('CODE'))
    - Supports mixing database timeseries with dynamically computed series

    Examples:
    - ["AAPL", "MSFT"] - database codes
    - ["AAPL", "Series('SPY')"] - mix of database and expression
    - ["Series('AAPL').pct_change()"] - pure expression

    Note: For codes containing commas or special characters, use JSON body or JSON in X-Codes header.

    Returns column-oriented format with dates sorted descending.
    """
    ensure_connection()

    try:
        codes = await _parse_codes_from_request(request)
        logger.debug("Processing %d codes: %s", len(codes), codes)

        def _process_all_codes_sync():
            series_list = []
            for code in codes:
                try:
                    ts = (
                        db.query(Timeseries)
                        .options(joinedload(Timeseries.data_record))
                        .filter(Timeseries.code == code)
                        .first()
                    )

                    if ts:
                        ts_data = _process_database_timeseries(ts, db)
                        if ts_data is not None:
                            series_list.append(ts_data)
                    else:
                        evaluated_series = _evaluate_expression(
                            code, start_date, end_date
                        )
                        series_list.extend(evaluated_series)
                except Exception as e:
                    logger.warning("Error processing custom timeseries %s: %s", code, e)
                    continue

            if series_list:
                df = pd.concat(series_list, axis=1)
                return _format_dataframe_response(df, series_list, start_date, end_date)
            else:
                return Response(
                    content=json.dumps({"Date": []}, ensure_ascii=False),
                    media_type="application/json",
                )

        import asyncio

        return await asyncio.to_thread(_process_all_codes_sync)

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error retrieving custom timeseries: %s", e)
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/timeseries.exec")
@_limiter.limit("20/minute")
async def exec_code_block(
    request: Request,
    _current_user: User = Depends(get_current_user),
):
    """
    POST /api/timeseries.exec — Execute a multi-line code block that produces
    a DataFrame or Series.  The code must assign its output to ``result``.

    Body: ``{ "code": "..." }``

    Example:
        code = '''
        spy = Series("SPY US EQUITY:PX_LAST")
        qqq = Series("QQQ US EQUITY:PX_LAST")
        result = pd.concat([spy, qqq], axis=1)
        '''

    Returns column-oriented JSON ``{ "Date": [...], "col1": [...], ... }``.
    """
    try:
        body = await request.json()
        code = body.get("code", "")
        if not code or not code.strip():
            raise HTTPException(status_code=400, detail="Empty code block")

        import asyncio

        def _run():
            evaluated = safe_exec_code(code, TIMESERIES_EXPRESSION_CONTEXT)
            if isinstance(evaluated, pd.Series):
                df = evaluated.to_frame()
            else:
                df = evaluated
            df.index.name = "Date"
            df = df.sort_index()
            # Format index — handle both datetime and non-datetime indices
            idx = df.index
            if hasattr(idx, 'strftime'):
                dates = [d.strftime("%Y-%m-%d") for d in idx]
            else:
                try:
                    dates = [pd.Timestamp(d).strftime("%Y-%m-%d") for d in idx]
                except Exception:
                    dates = [str(d) for d in idx]
            # Column-oriented JSON (same format as timeseries.custom)
            out: dict = {"Date": dates}
            for col in df.columns:
                vals = df[col].tolist()
                out[str(col)] = [
                    None if (v is None or (isinstance(v, float) and (pd.isna(v) or not math.isfinite(v)))) else v
                    for v in vals
                ]
            out["__columns__"] = [str(c) for c in df.columns]
            return out

        result = await asyncio.to_thread(_run)
        return result

    except UnsafeExpressionError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error executing code block: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ==============================================================================
# Market.xlsm Download
# ==============================================================================
import pathlib as _pathlib

_MARKET_FILE = _pathlib.Path(__file__).resolve().parents[4] / "Market.xlsm"


@router.get("/download/market")
def download_market_file(
    _current_user: User = Depends(get_current_user),
):
    """Download Market.xlsm Excel macro workbook."""
    if not _MARKET_FILE.is_file():
        raise HTTPException(status_code=404, detail="Market.xlsm not found on server")
    return FileResponse(
        path=str(_MARKET_FILE),
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        filename="Market.xlsm",
    )


def _merge_columnar_to_db(df: pd.DataFrame, db: SessionType) -> dict:
    """Merge a date-indexed DataFrame (columns=codes) into the database.
    Returns {"updated": [...], "not_found": [...], "points": int}.
    Uses batch loading to minimize DB queries.
    """
    from ix.db.models import TimeseriesData

    codes_list = list(df.columns)
    if not codes_list:
        return {"updated": [], "not_found": [], "points": 0}

    # Batch-load all matching Timeseries in one query
    all_ts = db.query(Timeseries).filter(Timeseries.code.in_(codes_list)).all()
    ts_by_code = {ts.code: ts for ts in all_ts}
    found_codes = set(ts_by_code.keys())
    not_found_codes = [c for c in codes_list if c not in found_codes]

    if not ts_by_code:
        return {"updated": [], "not_found": not_found_codes, "points": 0}

    # Batch-load all data records in one query
    ts_ids = [ts.id for ts in all_ts]
    all_data = db.query(TimeseriesData).filter(TimeseriesData.timeseries_id.in_(ts_ids)).all()
    data_by_ts_id = {dr.timeseries_id: dr for dr in all_data}

    updated_codes = []
    total_points = 0
    now = datetime.now()

    for code in codes_list:
        ts = ts_by_code.get(code)
        if ts is None:
            continue

        # Get or create data record (no extra query — already loaded)
        data_record = data_by_ts_id.get(ts.id)
        if data_record is None:
            data_record = TimeseriesData(timeseries_id=ts.id, data={})
            db.add(data_record)
            data_by_ts_id[ts.id] = data_record

        column_data = data_record.data if data_record.data else {}

        if column_data and isinstance(column_data, dict):
            existing_data = pd.Series(column_data)
            if not existing_data.empty:
                existing_data.index = pd.to_datetime(existing_data.index, errors="coerce")
                existing_data = existing_data.dropna()
        else:
            existing_data = pd.Series(dtype=float)

        new_series = df[code].dropna()

        if not existing_data.empty:
            combined = pd.concat([existing_data, new_series], axis=0)
            combined = combined[~combined.index.duplicated(keep="last")].sort_index()
        else:
            combined = new_series.sort_index()

        data_dict = {}
        for k, v in combined.to_dict().items():
            date_str = str(k.date()) if hasattr(k, "date") else str(pd.to_datetime(k).date()) if not isinstance(k, str) else k
            data_dict[date_str] = float(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else None

        data_record.data = data_dict
        data_record.updated = now
        ts.start = combined.index.min().date() if len(combined) > 0 else None
        ts.end = combined.index.max().date() if len(combined) > 0 else None
        ts.num_data = len(combined)
        ts.updated = now

        total_points += len(new_series)
        updated_codes.append(code)

    db.commit()
    return {"updated": updated_codes, "not_found": not_found_codes, "points": total_points}


def _sync_to_cloud(df: pd.DataFrame, response: dict) -> None:
    """Sync a DataFrame to the cloud (Railway) database. Non-fatal on failure."""
    from ix.db.conn import cloud_session

    try:
        with cloud_session() as cloud_db:
            if cloud_db is None:
                return
            cloud_result = _merge_columnar_to_db(df, cloud_db)
            response["cloud_updated"] = cloud_result["updated"]
            response["cloud_points_merged"] = cloud_result["points"]
            if cloud_result["not_found"]:
                response.setdefault("warning", "")
                if response["warning"]:
                    response["warning"] += "; "
                response["warning"] += f"Cloud DB codes not found: {cloud_result['not_found']}"
            logger.info(
                "Cloud DB: synced %d codes (%d points)",
                len(cloud_result["updated"]), cloud_result["points"],
            )
    except Exception as e:
        logger.exception("Cloud DB sync failed: %s", e)
        response["cloud_warning"] = f"Cloud sync failed: {str(e)}"


@router.post("/upload_data")
@_limiter.limit("10/minute")
def upload_data(
    request: Request,
    payload: TimeseriesDataUpload,
    db: SessionType = Depends(get_db),
    _current_user: User = Depends(get_current_admin_user),
):
    """
    POST /api/upload_data - Upload timeseries data.
    Server: saves JSON to R2.
    Local: merges directly into local DB + cloud DB (no R2).
    """
    from ix.db.boto import Boto
    from ix.misc.settings import Settings

    if not payload.data:
        raise HTTPException(status_code=400, detail="No data provided")

    records = [{"date": d.date, "code": d.code, "value": d.value} for d in payload.data]
    codes = sorted(set(d.code for d in payload.data))

    # Server: save to R2 only
    if Settings.is_server:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"uploads/{timestamp}_row_{len(records)}rec.json"
        try:
            Boto().save_json({"format": "row", "records": records}, filename)
        except Exception as e:
            logger.exception("Failed to upload data to R2: %s", e)
            raise HTTPException(status_code=500, detail="Failed to save to storage")
        return {"message": f"Saved {len(records)} records to storage.", "file": filename, "codes": codes}

    # Local: merge into local DB + cloud DB
    ensure_connection()
    df = pd.DataFrame(records)
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["value", "date", "code"])
    if df.empty:
        raise HTTPException(status_code=400, detail="No valid records after cleaning.")

    pivoted = df.pivot(index="date", columns="code", values="value").sort_index()
    pivoted = pivoted.dropna(how="all", axis=1).dropna(how="all", axis=0)
    pivoted.index = pd.to_datetime(pivoted.index)

    result = _merge_columnar_to_db(pivoted, db)
    response = {"message": f"Merged {result['points']} points for {len(result['updated'])} codes.", "db_updated": result["updated"], "db_points_merged": result["points"]}
    if result["not_found"]:
        response["warning"] = f"Codes not found in database: {result['not_found']}"

    _sync_to_cloud(pivoted, response)
    return response


@router.post("/upload_data_columnar")
@_limiter.limit("10/minute")
def upload_data_columnar(
    request: Request,
    payload: TimeseriesColumnarUpload,
    db: SessionType = Depends(get_db),
    _current_user: User = Depends(get_current_admin_user),
):
    """
    POST /api/upload_data_columnar - Upload timeseries data.
    Server: saves JSON to R2.
    Local: merges directly into local DB + cloud DB (no R2).
    """
    from ix.db.boto import Boto
    from ix.misc.settings import Settings

    if not payload.dates or not payload.columns:
        raise HTTPException(status_code=400, detail="No data provided")

    num_dates = len(payload.dates)
    num_cols = len(payload.columns)
    codes = sorted(payload.columns.keys())

    # Server: save to R2 only
    if Settings.is_server:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"uploads/{timestamp}_col_{num_cols}x{num_dates}.json"
        try:
            Boto().save_json(
                {"format": "columnar", "dates": payload.dates, "columns": payload.columns},
                filename,
            )
        except Exception as e:
            logger.exception("Failed to upload columnar data to R2: %s", e)
            raise HTTPException(status_code=500, detail="Failed to save to storage")
        return {"message": f"Saved {num_dates} dates × {num_cols} columns to storage.", "file": filename, "codes": codes}

    # Local: merge into local DB + cloud DB
    ensure_connection()
    dates_index = pd.to_datetime(payload.dates, errors="coerce")
    df = pd.DataFrame(payload.columns, index=dates_index)
    df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)
    if df.empty:
        raise HTTPException(status_code=400, detail="No valid records after cleaning.")

    result = _merge_columnar_to_db(df, db)
    response = {"message": f"Merged {result['points']} points for {len(result['updated'])} codes.", "db_updated": result["updated"], "db_points_merged": result["points"]}
    if result["not_found"]:
        response["warning"] = f"Codes not found in database: {result['not_found']}"

    _sync_to_cloud(df, response)
    return response


@router.post("/sync_uploads")
@_limiter.limit("10/minute")
def sync_uploads_from_r2(
    request: Request,
    db: SessionType = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """
    POST /api/sync_uploads - (Local only) Process pending R2 upload files
    into local DB + cloud DB, then move them to uploads/processed/.
    """
    from ix.db.boto import Boto
    from ix.misc.settings import Settings

    if Settings.is_server:
        raise HTTPException(status_code=400, detail="Sync is only available on local env.")

    try:
        ensure_connection()
        storage = Boto()
        pending = storage.list_prefix("uploads/")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"R2/DB initialization failed: {e}")

    # Exclude processed, failed, and retry-counter metadata files
    pending = [k for k in pending if
        not k.startswith("uploads/processed/")
        and not k.startswith("uploads/failed/")
        and not k.endswith(".retries")]

    if not pending:
        return {"message": "No pending uploads.", "processed": 0}

    results = []
    for key in sorted(pending):
        try:
            data = storage.get_json(key)
            if not data:
                logger.warning("Empty or invalid JSON: %s", key)
                continue

            fmt = data.get("format", "")
            if fmt == "columnar":
                dates_index = pd.to_datetime(data["dates"], errors="coerce")
                df = pd.DataFrame(data["columns"], index=dates_index)
            elif fmt == "row":
                records = data.get("records", [])
                raw = pd.DataFrame(records)
                raw["value"] = pd.to_numeric(raw["value"], errors="coerce")
                raw["date"] = pd.to_datetime(raw["date"], errors="coerce")
                raw = raw.dropna(subset=["value", "date", "code"])
                df = raw.pivot(index="date", columns="code", values="value").sort_index()
                df.index = pd.to_datetime(df.index)
            else:
                logger.warning("Unknown format in %s: %s", key, fmt)
                continue

            df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)
            if df.empty:
                logger.info("No data after cleaning in %s", key)
                storage.rename_file(key, key.replace("uploads/", "uploads/processed/", 1))
                continue

            # Merge into local DB
            result = _merge_columnar_to_db(df, db)
            logger.info("Synced %s: %d codes, %d points", key, len(result["updated"]), result["points"])

            # Sync to cloud DB
            file_response = {}
            _sync_to_cloud(df, file_response)

            # Move to processed
            storage.rename_file(key, key.replace("uploads/", "uploads/processed/", 1))

            results.append({
                "file": key,
                "codes": result["updated"],
                "points": result["points"],
                "cloud": file_response.get("cloud_updated", []),
            })
        except Exception as e:
            logger.exception("Failed to process %s: %s", key, e)
            results.append({"file": key, "error": str(e)})

    return {"message": f"Processed {len(results)} file(s).", "processed": len(results), "details": results}
