from dash import dcc, html, callback, Output, Input, State
import dash_bootstrap_components as dbc
import pandas as pd
import io, base64
from ix.misc.email import EmailSender
from ix.misc.settings import Settings
from ix.misc.terminal import get_logger

logger = get_logger(__name__)


# Layout wrapped in a Card for a polished look.
layout = dbc.Container(
    fluid=True,
    style={
        "backgroundColor": "transparent",
        "color": "#f8f9fa",
        "padding": "10px",
    },
    children=[
        dbc.Card(
            className="shadow rounded-3 w-100",
            style={
                "backgroundColor": "transparent",
                "border": "1px solid #f8f9fa",
                "boxShadow": "2px 2px 5px rgba(0,0,0,0.5)",
                "marginBottom": "1rem",
            },
            children=[
                dbc.CardHeader(
                    html.H3("Excel File Uploader", className="mb-0"),
                    style={
                        "backgroundColor": "transparent",
                        "color": "#f8f9fa",
                        "borderBottom": "2px solid #f8f9fa",
                        "padding": "1rem",
                    },
                ),
                dbc.CardBody(
                    style={
                        "backgroundColor": "transparent",
                        "color": "#f8f9fa",
                        "padding": "1.5rem",
                    },
                    children=[
                        dcc.Upload(
                            id="upload-data",
                            children=html.Div(
                                ["Drag and Drop or ", html.A("Select Files")]
                            ),
                            style={
                                "width": "100%",
                                "height": "60px",
                                "lineHeight": "60px",
                                "borderWidth": "1px",
                                "borderStyle": "dashed",
                                "borderRadius": "5px",
                                "textAlign": "center",
                                "margin": "10px 0",
                                "backgroundColor": "transparent",
                                "color": "#f8f9fa",
                            },
                            multiple=False,
                        ),
                        html.Div(
                            id="output-message",
                            style={"marginTop": "10px"},
                        ),
                    ],
                ),
            ],
        ),
    ],
)


from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


@callback(
    Output("output-message", "children"),
    Input("upload-data", "contents"),
    State("upload-data", "filename"),
)
def update_output(contents, filename):
    """
    Process the uploaded Excel file by reading its 'Data' sheet,
    uploading data to Bloomberg, and adding a new sheet 'DataV'
    to the original workbook. The modified workbook is then emailed.
    """
    if not contents:
        return ""

    try:
        # Decode the base64 content and create an in-memory file
        header, content_string = contents.split(",", 1)
        decoded = base64.b64decode(content_string)
        in_file = io.BytesIO(decoded)

        # Read the 'Data' sheet using pandas and process it
        data = pd.read_excel(in_file, sheet_name="Data", parse_dates=True, index_col=[0])
        upload_bbg_data(data=data)

        # Reset pointer to the beginning for openpyxl to load the workbook
        in_file.seek(0)
        wb = load_workbook(in_file)

        # If "DataV" exists, remove it (or you could update it as needed)
        if "DataV" in wb.sheetnames:
            ws = wb["DataV"]
            wb.remove(ws)

        # Create a new sheet "DataV"
        ws = wb.create_sheet("DataV")
        # Write the DataFrame data to the new sheet
        for row in dataframe_to_rows(data, index=False, header=True):
            ws.append(row)

        # Save the updated workbook back to the original in-memory file
        in_file.seek(0)
        in_file.truncate()  # Clear existing content
        wb.save(in_file)
        in_file.seek(0)  # Reset pointer before sending

        # Prepare and send the email with the updated Excel file attached
        email_sender = EmailSender(
            to=", ".join(Settings.email_recipients),
            subject="[IX] Daily Data Share",
            content="Please find the attached Excel file with the latest data and added sheet.",
        )
        email_sender.attach(in_file, filename=filename)
        email_sender.send()

        logger.info(
            f"Email sent successfully to {', '.join(Settings.email_recipients)}"
        )
        return html.Div(
            [html.P(f"File '{filename}' uploaded successfully, updated, and emailed!")]
        )
    except Exception as e:
        logger.error(f"Failed to process and send email: {str(e)}", exc_info=True)
        return html.Div(f"Error processing file: {str(e)}")


from ix.db.models import Metadata
def upload_bbg_data(data: pd.DataFrame) -> bool:
    for ticker_field in data.columns:
        ticker, field = str(ticker_field).split(":", maxsplit=2)
        metadata = Metadata.find_one({"bbg_ticker": ticker}).run()
        if metadata:
            ts = data[ticker_field].dropna()
            if ts.empty:
                continue
            metadata.ts(field=field).data = ts
    return True
